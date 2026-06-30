"""
=============================================================================
MODELO DE RÉGIMEN RISK-ON / RISK-OFF PARA MERVAL USD - CÓDIGO DE PRODUCCIÓN
=============================================================================
Pipeline integrado con:
    - Ensemble de 5 señales equal-weighted
    - Walk-forward refit del Markov Switching cada 24 meses
    - Filtered probabilities (sin lookahead)
    - Histéresis asimétrica 0.55/0.45, confirm 2m, cooldown 1m
    - Sin apalancamiento (decisión binaria 0/100)

OUTPUTS:
    - Logging detallado en consola y archivo de log
    - Excel multi-sheet con backtest completo
    - Gráficos múltiples en PNG

USO:
    python production_model.py
=============================================================================
"""

from __future__ import annotations
import os
import sys
import logging
import warnings
from dataclasses import dataclass, field, asdict
from datetime import datetime
from typing import Optional

import numpy as np
import pandas as pd
import matplotlib.pyplot as plt
import matplotlib.dates as mdates
from matplotlib.patches import Patch
import seaborn as sns

from statsmodels.tsa.regime_switching.markov_regression import MarkovRegression
from scipy import stats
from sklearn.metrics import (
    f1_score, precision_score, recall_score, accuracy_score,
    confusion_matrix, cohen_kappa_score, roc_auc_score, classification_report
)

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import ColorScaleRule, CellIsRule
from openpyxl.styles.differential import DifferentialStyle
from openpyxl.formatting.rule import Rule

warnings.filterwarnings("ignore")

# =============================================================================
# CONFIGURACIÓN GLOBAL
# =============================================================================

@dataclass
class Config:
    """Configuración completa del modelo de producción."""

    # Paths
    input_file: str = "/mnt/user-data/uploads/Merval_RP.xlsx"
    output_dir: str = "/mnt/user-data/outputs/production"
    log_file: str = "production_run.log"
    excel_output: str = "Merval_Regime_Backtest.xlsx"

    # Column names (case insensitive matching)
    col_date: str = "DATE"
    col_pe: str = "P/E E Trend"
    col_rp: str = "Riesgo pais"
    col_merval: str = "Merval USD"

    # Ensemble - composite weights
    composite_w_rp: float = 0.40
    composite_w_pe: float = 0.30
    composite_w_drp: float = 0.30
    composite_window: int = 36
    composite_k_sigmoid: float = 1.2

    # Signal 3 - RP Trend
    rp_trend_window: int = 6
    rp_trend_smooth: int = 3

    # Signal 4 - Price Momentum
    momentum_lookback: int = 6

    # Signal 5 - Spread Acceleration
    spread_smooth: int = 3
    spread_lookback: int = 6
    spread_norm_window: int = 36

    # Markov Switching
    ms_n_regimes: int = 2
    ms_search_reps: int = 20
    ms_seed: int = 42
    ms_refit_months: int = 24
    ms_initial_train: int = 72

    # Decision logic - thresholds
    theta_in: float = 0.55
    theta_out: float = 0.45
    confirm_k: int = 2
    cooldown: int = 1

    # Triple Barrier (for evaluation only)
    tb_upper_pct: float = 0.20
    tb_lower_pct: float = 0.20
    tb_horizon: int = 12

    # Backtest
    tc_bps: float = 30.0
    initial_capital: float = 100.0

    # Plotting
    plot_dpi: int = 110
    plot_style: str = "seaborn-v0_8-darkgrid"

    def to_dict(self) -> dict:
        return asdict(self)


CONFIG = Config()


# =============================================================================
# LOGGING SETUP
# =============================================================================

def setup_logger(log_path: str) -> logging.Logger:
    """Configura logger con output a consola y archivo."""
    logger = logging.getLogger("MervalRegime")
    logger.setLevel(logging.DEBUG)
    logger.handlers.clear()

    formatter = logging.Formatter(
        "%(asctime)s | %(levelname)-7s | %(message)s",
        datefmt="%Y-%m-%d %H:%M:%S"
    )

    console = logging.StreamHandler(sys.stdout)
    console.setLevel(logging.INFO)
    console.setFormatter(formatter)
    logger.addHandler(console)

    file_handler = logging.FileHandler(log_path, mode="w", encoding="utf-8")
    file_handler.setLevel(logging.DEBUG)
    file_handler.setFormatter(formatter)
    logger.addHandler(file_handler)

    return logger


def log_banner(logger: logging.Logger, text: str, char: str = "=", width: int = 80):
    logger.info(char * width)
    logger.info(text.center(width))
    logger.info(char * width)


# =============================================================================
# UTILITY FUNCTIONS
# =============================================================================

def sigmoid(x, k: float = 1.0):
    return 1 / (1 + np.exp(-k * np.asarray(x, dtype=float)))


def robust_zscore(s: pd.Series, window: int, min_periods: Optional[int] = None) -> pd.Series:
    """z-score robusto: (x - mediana) / (1.4826 * MAD), ambos rolling."""
    min_periods = min_periods or window // 2
    med = s.rolling(window, min_periods=min_periods).median()
    mad = (s - med).abs().rolling(window, min_periods=min_periods).median()
    return (s - med) / (1.4826 * mad.replace(0, np.nan))


# =============================================================================
# 1. DATA LOADING
# =============================================================================

def load_inputs(cfg: Config, logger: logging.Logger) -> pd.DataFrame:
    """Carga y valida los inputs del Excel."""
    log_banner(logger, "PASO 1: CARGA DE INPUTS")
    logger.info(f"Leyendo archivo: {cfg.input_file}")

    if not os.path.exists(cfg.input_file):
        raise FileNotFoundError(f"Archivo de input no encontrado: {cfg.input_file}")

    df = pd.read_excel(cfg.input_file)
    df.columns = [c.strip() for c in df.columns]

    # Parse de fechas (soporta serial Excel o datetime)
    if pd.api.types.is_numeric_dtype(df[cfg.col_date]):
        df[cfg.col_date] = pd.to_datetime(df[cfg.col_date], origin="1899-12-30", unit="D")
    else:
        df[cfg.col_date] = pd.to_datetime(df[cfg.col_date])

    df = df.rename(columns={
        cfg.col_pe: "pe_trend",
        cfg.col_rp: "rp",
        cfg.col_merval: "merval",
    })
    df = df.set_index(cfg.col_date).sort_index()
    df["ret"] = np.log(df["merval"]).diff()

    logger.info(f"Período de datos: {df.index.min().date()} → {df.index.max().date()}")
    logger.info(f"Observaciones totales: {len(df)}")
    logger.info(f"Observaciones con retorno válido: {df['ret'].notna().sum()}")

    logger.debug("\nEstadísticas descriptivas:")
    logger.debug(df[["pe_trend", "rp", "merval", "ret"]].describe().round(4).to_string())

    # Validación de integridad
    if df["rp"].max() > 1.0:
        logger.warning(f"Riesgo País máximo > 100% ({df['rp'].max():.4f}). "
                       f"Confirmar si está en bps/10000 o en decimal.")
    if df["merval"].min() <= 0:
        raise ValueError("Merval con valor <= 0; no se puede calcular log returns.")

    return df


# =============================================================================
# 2. SIGNAL GENERATION
# =============================================================================

def fit_ms_window(rets: pd.Series, cfg: Config, logger: logging.Logger,
                  context: str = "") -> dict:
    """Ajusta MS-AR(1) y devuelve filtered prob + params."""
    y = rets.dropna()
    np.random.seed(cfg.ms_seed)
    model = MarkovRegression(
        y.values, k_regimes=cfg.ms_n_regimes,
        trend="c", switching_variance=True
    )
    res = model.fit(disp=False, search_reps=cfg.ms_search_reps)

    smoothed = np.asarray(res.smoothed_marginal_probabilities)
    if smoothed.shape[0] == cfg.ms_n_regimes:
        smoothed = smoothed.T
    cond_means = [
        np.sum(smoothed[:, k] * y.values) / np.sum(smoothed[:, k])
        for k in range(cfg.ms_n_regimes)
    ]
    risk_on_idx = int(np.argmax(cond_means))

    filtered = np.asarray(res.filtered_marginal_probabilities)
    if filtered.shape[0] == cfg.ms_n_regimes:
        filtered = filtered.T

    p_filt = pd.Series(filtered[:, risk_on_idx], index=y.index, name="p_ms")
    p_smooth = pd.Series(smoothed[:, risk_on_idx], index=y.index, name="p_ms_smooth")

    if context:
        logger.debug(f"  MS fit [{context}]: "
                     f"μ_on={cond_means[risk_on_idx]*12:.3f}, "
                     f"μ_off={cond_means[1-risk_on_idx]*12:.3f}, "
                     f"N={len(y)}")

    return {
        "filtered": p_filt, "smoothed": p_smooth,
        "cond_means_annual": [m * 12 for m in cond_means],
        "risk_on_idx": risk_on_idx, "n_obs": len(y), "params": res.params,
    }


def signal_composite(df: pd.DataFrame, cfg: Config) -> pd.Series:
    """Señal 2: Composite z-score robusto."""
    z_rp = robust_zscore(df["rp"], cfg.composite_window)
    z_pe = robust_zscore(df["pe_trend"], cfg.composite_window)
    drp_12m = df["rp"].diff(12)
    z_drp = robust_zscore(drp_12m, cfg.composite_window)
    composite = (cfg.composite_w_rp * (-z_rp)
                 + cfg.composite_w_pe * z_pe
                 + cfg.composite_w_drp * (-z_drp))
    composite = composite.clip(-3, 3)
    return pd.Series(sigmoid(composite, k=cfg.composite_k_sigmoid),
                     index=composite.index, name="p_composite")


def signal_rp_trend(rp: pd.Series, cfg: Config) -> pd.Series:
    """Señal 3: Tendencia suavizada del RP."""
    rp_smooth = rp.ewm(span=cfg.rp_trend_smooth, adjust=False).mean()
    drp = rp_smooth.diff(cfg.rp_trend_window)
    sigma = drp.rolling(36, min_periods=18).std()
    z = (-drp / sigma).clip(-3, 3)
    return pd.Series(sigmoid(z, k=1.2), index=z.index, name="p_rp_trend")


def signal_momentum(rets: pd.Series, cfg: Config) -> pd.Series:
    """Señal 4: Momentum del Merval USD."""
    mom = rets.rolling(cfg.momentum_lookback).sum()
    vol = rets.rolling(cfg.momentum_lookback).std() * np.sqrt(cfg.momentum_lookback)
    z = (mom / vol).clip(-3, 3)
    return pd.Series(sigmoid(z, k=1.0), index=z.index, name="p_momentum")


def signal_spread_accel(rp: pd.Series, cfg: Config) -> pd.Series:
    """Señal 5: Spread Acceleration (Δ²RP)."""
    rp_smooth = rp.ewm(span=cfg.spread_smooth, adjust=False).mean()
    velocity = rp_smooth.diff(cfg.spread_lookback)
    acceleration = velocity.diff(cfg.spread_lookback)
    sigma = acceleration.rolling(cfg.spread_norm_window, min_periods=18).std()
    z = (-acceleration / sigma).clip(-3, 3)
    return pd.Series(sigmoid(z, k=1.0), index=z.index, name="p_spread_accel")


def build_all_signals(df: pd.DataFrame, cfg: Config, logger: logging.Logger,
                       ms_filtered: pd.Series) -> pd.DataFrame:
    """Construye las 5 señales del ensemble."""
    log_banner(logger, "PASO 2: CONSTRUCCIÓN DE SEÑALES")
    rets = df["ret"]

    logger.info("Generando Señal 1: Markov Switching (filtered probabilities)...")
    s1 = ms_filtered

    logger.info("Generando Señal 2: Composite z-score robusto...")
    s2 = signal_composite(df, cfg)

    logger.info("Generando Señal 3: RP Trend...")
    s3 = signal_rp_trend(df["rp"], cfg)

    logger.info("Generando Señal 4: Price Momentum...")
    s4 = signal_momentum(rets, cfg)

    logger.info("Generando Señal 5: Spread Acceleration (Δ²RP)...")
    s5 = signal_spread_accel(df["rp"], cfg)

    signals_df = pd.concat([s1, s2, s3, s4, s5], axis=1)
    logger.info(f"Señales construidas. Shape: {signals_df.shape}")

    logger.debug("\nEstadísticas de las señales:")
    logger.debug(signals_df.describe().round(4).to_string())

    return signals_df


# =============================================================================
# 3. WALK-FORWARD REFIT DEL MS
# =============================================================================

def walkforward_ms_refit(rets: pd.Series, cfg: Config,
                          logger: logging.Logger) -> pd.Series:
    """Genera la serie completa de filtered prob de MS con refit cada N meses."""
    log_banner(logger, "PASO 3: WALK-FORWARD REFIT DEL MARKOV SWITCHING")

    valid_idx = rets.dropna().index
    n = len(valid_idx)

    if n < cfg.ms_initial_train:
        logger.warning(f"Insuficientes datos para WF refit (N={n} < initial_train={cfg.ms_initial_train}). "
                       f"Usando fit único sobre toda la serie.")
        ms_fit = fit_ms_window(rets, cfg, logger, "fit único")
        return ms_fit["filtered"]

    # Splits: refit cada `ms_refit_months` meses desde initial_train
    splits = []
    start = cfg.ms_initial_train
    while start < n:
        end = min(start + cfg.ms_refit_months, n)
        splits.append({"train_end": start, "valid_end": end})
        start = end

    logger.info(f"Splits programados: {len(splits)}")
    logger.info(f"Initial train: {cfg.ms_initial_train} meses ({valid_idx[0].date()} → "
                f"{valid_idx[cfg.ms_initial_train-1].date()})")
    logger.info(f"Refit cada: {cfg.ms_refit_months} meses")

    # Para cada split: refit con datos hasta valid_end, usar filtered probs
    # del segmento (train_end, valid_end]
    full_prob = pd.Series(index=valid_idx, dtype=float, name="p_ms")
    refit_log = []

    # Segmento inicial (initial train period): usar fit sobre initial_train
    logger.info(f"\nFit del período inicial (warm-up):")
    initial_window = rets.loc[valid_idx[:cfg.ms_initial_train]]
    fit0 = fit_ms_window(initial_window, cfg, logger,
                          f"initial t≤{valid_idx[cfg.ms_initial_train-1].date()}")
    full_prob.loc[valid_idx[:cfg.ms_initial_train]] = fit0["filtered"].values
    refit_log.append({
        "split": 0,
        "train_end_date": valid_idx[cfg.ms_initial_train-1].date(),
        "mu_on_annual": fit0["cond_means_annual"][fit0["risk_on_idx"]],
        "mu_off_annual": fit0["cond_means_annual"][1-fit0["risk_on_idx"]],
        "n_obs": fit0["n_obs"],
    })

    # Splits subsiguientes
    logger.info(f"\nFits walk-forward:")
    for i, sp in enumerate(splits, 1):
        train_end_date = valid_idx[sp["train_end"]-1]
        valid_end_date = valid_idx[sp["valid_end"]-1]
        # Refit expanding window hasta valid_end
        window = rets.loc[valid_idx[:sp["valid_end"]]]
        fit = fit_ms_window(window, cfg, logger,
                             f"split {i} t≤{valid_end_date.date()}")
        # Asignar el segmento OOS (train_end+1 a valid_end)
        oos_dates = valid_idx[sp["train_end"]:sp["valid_end"]]
        oos_prob = fit["filtered"].loc[oos_dates]
        full_prob.loc[oos_dates] = oos_prob.values
        refit_log.append({
            "split": i,
            "train_end_date": train_end_date.date(),
            "valid_end_date": valid_end_date.date(),
            "mu_on_annual": fit["cond_means_annual"][fit["risk_on_idx"]],
            "mu_off_annual": fit["cond_means_annual"][1-fit["risk_on_idx"]],
            "n_obs": fit["n_obs"],
        })

    # Resumen del refit
    refit_df = pd.DataFrame(refit_log)
    logger.info(f"\nResumen de parámetros MS por refit:")
    logger.info(refit_df.to_string(index=False))

    return full_prob


# =============================================================================
# 4. ENSEMBLE Y DECISIÓN
# =============================================================================

def build_ensemble(signals_df: pd.DataFrame, logger: logging.Logger) -> pd.Series:
    """Construye la prob agregada del ensemble (equal-weighted)."""
    log_banner(logger, "PASO 4: AGREGACIÓN DEL ENSEMBLE")
    valid = signals_df.dropna()
    logger.info(f"Períodos con las 5 señales válidas: {len(valid)} de {len(signals_df)}")
    logger.info(f"Período efectivo del ensemble: {valid.index.min().date()} → {valid.index.max().date()}")

    p_ensemble = valid.mean(axis=1).rename("p_ensemble")

    logger.info(f"\nDistribución de la prob agregada:")
    logger.info(p_ensemble.describe().round(4).to_string())

    return p_ensemble


def apply_hysteresis(p_ensemble: pd.Series, cfg: Config,
                      logger: logging.Logger) -> pd.DataFrame:
    """Aplica histéresis asimétrica con confirmación temporal."""
    log_banner(logger, "PASO 5: APLICACIÓN DE LA LÓGICA DE DECISIÓN")
    logger.info(f"Parámetros:")
    logger.info(f"  θ_in (entrada Risk-On)  = {cfg.theta_in}")
    logger.info(f"  θ_out (salida Risk-Off) = {cfg.theta_out}")
    logger.info(f"  Confirmación entrada    = {cfg.confirm_k} meses consecutivos")
    logger.info(f"  Cooldown mínimo         = {cfg.cooldown} meses")

    p = p_ensemble.copy()
    raw_on = (p > cfg.theta_in)
    raw_off = (p < cfg.theta_out)
    confirmed_on = raw_on.rolling(cfg.confirm_k).sum() >= cfg.confirm_k

    regime = pd.Series(0.0, index=p.index, name="regime")
    last_switch = 0
    switches_log = []
    for i in range(1, len(p)):
        prev = regime.iloc[i-1]
        since = i - last_switch
        new_regime = prev
        action = "hold"
        if prev == 0:
            if confirmed_on.iloc[i] and since >= cfg.cooldown:
                new_regime = 1
                action = "ENTER_RiskOn"
                last_switch = i
        else:
            if raw_off.iloc[i] and since >= cfg.cooldown:
                new_regime = 0
                action = "EXIT_to_RiskOff"
                last_switch = i
        regime.iloc[i] = new_regime
        if action != "hold":
            switches_log.append({
                "date": p.index[i], "action": action,
                "p_ensemble": p.iloc[i],
                "from_regime": "Risk-Off" if prev == 0 else "Risk-On",
                "to_regime": "Risk-On" if new_regime == 1 else "Risk-Off",
            })

    decision_df = pd.concat([p, regime], axis=1)
    decision_df["regime_label"] = decision_df["regime"].map({1: "Risk-On", 0: "Risk-Off"})

    n_switches = len(switches_log)
    pct_on = regime.mean()
    logger.info(f"\nResultados de la decisión:")
    logger.info(f"  # Switches totales: {n_switches}")
    logger.info(f"  % tiempo en Risk-On: {pct_on:.2%}")

    if switches_log:
        logger.info(f"\nLog completo de switches:")
        for sw in switches_log:
            logger.info(f"  {sw['date'].date()} | {sw['action']:18s} | "
                        f"p={sw['p_ensemble']:.3f} | {sw['from_regime']} → {sw['to_regime']}")

    return decision_df, pd.DataFrame(switches_log)


# =============================================================================
# 5. BACKTEST
# =============================================================================

def run_backtest(decision_df: pd.DataFrame, rets: pd.Series,
                  cfg: Config, logger: logging.Logger) -> pd.DataFrame:
    """Ejecuta el backtest binario con costos de transacción."""
    log_banner(logger, "PASO 6: BACKTEST BINARIO")
    logger.info(f"Capital inicial: ${cfg.initial_capital:,.2f}")
    logger.info(f"Costo de transacción: {cfg.tc_bps} bps por switch")

    bt = decision_df.copy()
    bt["ret_merval"] = rets
    bt["exposure"] = bt["regime"].shift(1).fillna(0)  # ejecución t+1
    bt["turnover"] = bt["exposure"].diff().abs().fillna(0)
    bt["tc_drag"] = bt["turnover"] * (cfg.tc_bps / 1e4)
    bt["ret_strat"] = bt["exposure"] * bt["ret_merval"] - bt["tc_drag"]
    bt["ret_bh"] = bt["ret_merval"]

    bt["equity_strat"] = cfg.initial_capital * (1 + bt["ret_strat"].fillna(0)).cumprod()
    bt["equity_bh"] = cfg.initial_capital * (1 + bt["ret_bh"].fillna(0)).cumprod()

    bt["peak_strat"] = bt["equity_strat"].cummax()
    bt["dd_strat"] = bt["equity_strat"] / bt["peak_strat"] - 1
    bt["peak_bh"] = bt["equity_bh"].cummax()
    bt["dd_bh"] = bt["equity_bh"] / bt["peak_bh"] - 1

    logger.info(f"\nEquity final:")
    logger.info(f"  Estrategia:   ${bt['equity_strat'].iloc[-1]:,.2f}  "
                f"({(bt['equity_strat'].iloc[-1]/cfg.initial_capital - 1):+.2%})")
    logger.info(f"  Buy & Hold:   ${bt['equity_bh'].iloc[-1]:,.2f}  "
                f"({(bt['equity_bh'].iloc[-1]/cfg.initial_capital - 1):+.2%})")
    logger.info(f"  Diferencial:  ${bt['equity_strat'].iloc[-1] - bt['equity_bh'].iloc[-1]:+,.2f}")
    logger.info(f"\n  Costos acumulados (TC): ${bt['tc_drag'].sum() * cfg.initial_capital:.2f}")

    return bt


# =============================================================================
# 6. MÉTRICAS DE PERFORMANCE
# =============================================================================

def perf_metrics(ret_series: pd.Series, freq: int = 12) -> dict:
    r = ret_series.dropna()
    if len(r) == 0:
        return {k: np.nan for k in ["CAGR", "Vol", "Sharpe", "Sortino", "MaxDD",
                                      "Calmar", "Skew", "ExKurt", "Obs"]}
    mu = r.mean() * freq
    sigma = r.std(ddof=1) * np.sqrt(freq)
    sharpe = mu / sigma if sigma > 0 else np.nan
    cagr = (1 + r).prod() ** (freq / len(r)) - 1 if len(r) > 0 else np.nan
    eq = (1 + r).cumprod()
    dd = eq / eq.cummax() - 1
    max_dd = dd.min()
    calmar = cagr / abs(max_dd) if max_dd < 0 else np.nan
    downside = r[r < 0].std(ddof=1) * np.sqrt(freq) if (r < 0).sum() > 1 else np.nan
    sortino = mu / downside if downside and downside > 0 else np.nan
    return {
        "CAGR": cagr, "Vol": sigma, "Sharpe": sharpe, "Sortino": sortino,
        "MaxDD": max_dd, "Calmar": calmar,
        "Skew": stats.skew(r), "ExKurt": stats.kurtosis(r),
        "Obs": len(r),
    }


def information_ratio(strat: pd.Series, bench: pd.Series, freq: int = 12) -> float:
    active = (strat - bench).dropna()
    if active.std() == 0 or len(active) == 0:
        return np.nan
    return (active.mean() / active.std()) * np.sqrt(freq)


def hit_ratios(bt: pd.DataFrame) -> dict:
    """Upside / downside capture y hit rates."""
    s = bt["ret_strat"].dropna()
    b = bt["ret_bh"].dropna()
    common = s.index.intersection(b.index)
    s, b = s.loc[common], b.loc[common]
    up_mask = b > 0
    dn_mask = b < 0
    uc = (s[up_mask].sum() / b[up_mask].sum()) if up_mask.sum() > 0 and b[up_mask].sum() != 0 else np.nan
    dc = (s[dn_mask].sum() / b[dn_mask].sum()) if dn_mask.sum() > 0 and b[dn_mask].sum() != 0 else np.nan
    win_rate = (s > 0).mean()
    return {"UpCapture": uc, "DownCapture": dc, "UC_DC_ratio": uc/dc if dc != 0 else np.nan,
            "WinRate": win_rate, "MonthsPositive": (s > 0).sum(), "MonthsNegative": (s < 0).sum()}


def yearly_returns(bt: pd.DataFrame) -> pd.DataFrame:
    bt2 = bt.copy()
    bt2["year"] = bt2.index.year
    out = bt2.groupby("year").agg(
        ret_strat=("ret_strat", lambda r: (1 + r).prod() - 1),
        ret_bh=("ret_bh", lambda r: (1 + r).prod() - 1),
        pct_on=("exposure", "mean"),
        n_switches=("turnover", lambda t: int(t.sum())),
    )
    out["alpha"] = out["ret_strat"] - out["ret_bh"]
    return out


def top_drawdowns(bt: pd.DataFrame, top_n: int = 10, col: str = "equity_strat") -> pd.DataFrame:
    """Identifica los N peores drawdowns con duraciones."""
    eq = bt[col]
    peak = eq.cummax()
    dd = eq / peak - 1
    is_dd = dd < 0
    drawdowns = []
    in_dd = False
    start = None
    for i in range(len(dd)):
        if is_dd.iloc[i] and not in_dd:
            in_dd = True
            start = i
        elif not is_dd.iloc[i] and in_dd:
            in_dd = False
            seg = dd.iloc[start:i+1]
            drawdowns.append({
                "start_date": dd.index[start].date(),
                "trough_date": seg.idxmin().date(),
                "end_date": dd.index[i].date(),
                "max_dd": seg.min(),
                "duration_months": i - start,
                "recovery_months": (dd.index[i] - seg.idxmin()).days // 30,
            })
    if in_dd:
        seg = dd.iloc[start:]
        drawdowns.append({
            "start_date": dd.index[start].date(),
            "trough_date": seg.idxmin().date(),
            "end_date": "ONGOING",
            "max_dd": seg.min(),
            "duration_months": len(dd) - start,
            "recovery_months": np.nan,
        })
    if not drawdowns:
        return pd.DataFrame()
    return pd.DataFrame(drawdowns).sort_values("max_dd").head(top_n).reset_index(drop=True)


def triple_barrier_labels(prices: pd.Series, cfg: Config) -> pd.Series:
    """Ground truth ex-post para evaluación clasificatoria."""
    n = len(prices)
    labels = pd.Series(index=prices.index, dtype="float64")
    for i in range(n - 1):
        p0 = prices.iloc[i]
        upper = p0 * (1 + cfg.tb_upper_pct)
        lower = p0 * (1 - cfg.tb_lower_pct)
        end = min(i + 1 + cfg.tb_horizon, n)
        future = prices.iloc[i+1:end]
        if len(future) == 0:
            continue
        hit_up = future > upper
        hit_dn = future < lower
        t_up = hit_up.idxmax() if hit_up.any() else None
        t_dn = hit_dn.idxmax() if hit_dn.any() else None
        if t_up is not None and t_dn is not None:
            labels.iloc[i] = 1.0 if t_up <= t_dn else 0.0
        elif t_up is not None:
            labels.iloc[i] = 1.0
        elif t_dn is not None:
            labels.iloc[i] = 0.0
        else:
            labels.iloc[i] = 1.0 if future.iloc[-1] > p0 else 0.0
    return labels.dropna()


def classification_eval(decision_df: pd.DataFrame, y_true: pd.Series) -> dict:
    common = decision_df.index.intersection(y_true.index)
    yp = decision_df.loc[common, "regime"].astype(int)
    yt = y_true.loc[common].astype(int)
    cm = confusion_matrix(yt, yp)
    out = {
        "Accuracy": accuracy_score(yt, yp),
        "Kappa": cohen_kappa_score(yt, yp),
        "F1_RiskOn": f1_score(yt, yp, pos_label=1, zero_division=0),
        "Precision_RiskOn": precision_score(yt, yp, pos_label=1, zero_division=0),
        "Recall_RiskOn": recall_score(yt, yp, pos_label=1, zero_division=0),
        "F1_RiskOff": f1_score(yt, yp, pos_label=0, zero_division=0),
        "Precision_RiskOff": precision_score(yt, yp, pos_label=0, zero_division=0),
        "Recall_RiskOff": recall_score(yt, yp, pos_label=0, zero_division=0),
        "TP_RiskOn": int(cm[1,1]), "FN_RiskOn": int(cm[1,0]),
        "TN_RiskOff": int(cm[0,0]), "FP_RiskOff": int(cm[0,1]),
        "N": len(yt),
    }
    try:
        out["AUC_ROC"] = roc_auc_score(yt, decision_df.loc[common, "p_ensemble"])
    except Exception:
        out["AUC_ROC"] = np.nan
    return out, cm


def consolidate_metrics(bt: pd.DataFrame, decision_df: pd.DataFrame,
                         y_true: pd.Series, logger: logging.Logger) -> dict:
    """Consolida todas las métricas en una estructura única."""
    log_banner(logger, "PASO 7: CÁLCULO DE MÉTRICAS DE PERFORMANCE")

    m_strat = perf_metrics(bt["ret_strat"])
    m_bh = perf_metrics(bt["ret_bh"])
    ir = information_ratio(bt["ret_strat"], bt["ret_bh"])
    hr = hit_ratios(bt)
    cls_metrics, cm = classification_eval(decision_df, y_true)
    yearly = yearly_returns(bt)
    top_dd = top_drawdowns(bt, top_n=10)

    n_switches = int(bt["turnover"].sum())
    pct_on = bt["exposure"].mean()

    summary = {
        "strategy": m_strat, "buy_hold": m_bh,
        "information_ratio": ir,
        "capture": hr,
        "classification": cls_metrics,
        "confusion_matrix": cm,
        "n_switches": n_switches, "pct_risk_on": pct_on,
        "yearly_returns": yearly,
        "top_drawdowns": top_dd,
    }

    # Impresión consolidada
    logger.info("\n" + "─" * 80)
    logger.info("PERFORMANCE CONSOLIDADA")
    logger.info("─" * 80)
    table = pd.DataFrame({"Estrategia": m_strat, "Buy & Hold": m_bh}).round(4)
    logger.info("\n" + table.to_string())
    logger.info(f"\nInformation Ratio vs B&H: {ir:.4f}")
    logger.info(f"Up Capture: {hr['UpCapture']:.2%} | Down Capture: {hr['DownCapture']:.2%} | "
                f"Ratio UC/DC: {hr['UC_DC_ratio']:.2f}")
    logger.info(f"Win Rate (months > 0): {hr['WinRate']:.2%}")

    logger.info("\n" + "─" * 80)
    logger.info("CLASIFICACIÓN vs GROUND TRUTH (Triple Barrier ex-post)")
    logger.info("─" * 80)
    for k, v in cls_metrics.items():
        if isinstance(v, float):
            logger.info(f"  {k:25s}: {v:.4f}")
        else:
            logger.info(f"  {k:25s}: {v}")

    logger.info("\n" + "─" * 80)
    logger.info("RETORNOS ANUALES")
    logger.info("─" * 80)
    logger.info("\n" + yearly.round(4).to_string())

    logger.info("\n" + "─" * 80)
    logger.info("TOP 10 DRAWDOWNS DE LA ESTRATEGIA")
    logger.info("─" * 80)
    logger.info("\n" + top_dd.to_string(index=False))

    return summary


# =============================================================================
# 7. VISUALIZACIONES
# =============================================================================

def generate_plots(bt: pd.DataFrame, decision_df: pd.DataFrame,
                    signals_df: pd.DataFrame, summary: dict,
                    cfg: Config, logger: logging.Logger) -> dict:
    """Genera todos los gráficos del reporte."""
    log_banner(logger, "PASO 8: GENERACIÓN DE GRÁFICOS")

    try:
        plt.style.use(cfg.plot_style)
    except OSError:
        plt.style.use("default")

    plot_paths = {}
    out = cfg.output_dir

    # ----- 1. Equity curves -----
    fig, ax = plt.subplots(figsize=(13, 6))
    ax.plot(bt.index, bt["equity_strat"], color="#1F4E79", linewidth=2.0,
            label=f"Estrategia (CAGR {summary['strategy']['CAGR']:.2%})")
    ax.plot(bt.index, bt["equity_bh"], color="#C00000", linewidth=1.5, linestyle="--",
            label=f"Buy & Hold (CAGR {summary['buy_hold']['CAGR']:.2%})")
    ax.set_yscale("log")
    ax.set_ylabel("Equity (base = $100, log scale)")
    ax.set_title("Equity Curves: Modelo de Régimen vs Buy & Hold", fontsize=13, fontweight="bold")
    ax.legend(loc="upper left", fontsize=10)
    ax.grid(True, alpha=0.3)
    ax.xaxis.set_major_locator(mdates.YearLocator(2))
    ax.xaxis.set_major_formatter(mdates.DateFormatter("%Y"))
    plt.tight_layout()
    p = f"{out}/01_equity_curves.png"
    plt.savefig(p, dpi=cfg.plot_dpi, bbox_inches="tight"); plt.close()
    plot_paths["equity"] = p

    # ----- 2. Drawdowns -----
    fig, ax = plt.subplots(figsize=(13, 5))
    ax.fill_between(bt.index, bt["dd_bh"], 0, color="#C00000", alpha=0.3, label="Buy & Hold")
    ax.fill_between(bt.index, bt["dd_strat"], 0, color="#1F4E79", alpha=0.5, label="Estrategia")
    ax.set_ylabel("Drawdown")
    ax.set_title("Drawdowns comparados", fontsize=13, fontweight="bold")
    ax.legend(loc="lower left", fontsize=10)
    ax.grid(True, alpha=0.3)
    ax.xaxis.set_major_locator(mdates.YearLocator(2))
    ax.xaxis.set_major_formatter(mdates.DateFormatter("%Y"))
    plt.tight_layout()
    p = f"{out}/02_drawdowns.png"
    plt.savefig(p, dpi=cfg.plot_dpi, bbox_inches="tight"); plt.close()
    plot_paths["drawdowns"] = p

    # ----- 3. Panel de señales -----
    fig, axes = plt.subplots(6, 1, figsize=(13, 14), sharex=True)
    sig_specs = [
        ("p_ms", "Señal 1: Markov Switching (filtered)", "#1f77b4"),
        ("p_composite", "Señal 2: Composite z-score", "#ff7f0e"),
        ("p_rp_trend", "Señal 3: RP Trend", "#2ca02c"),
        ("p_momentum", "Señal 4: Price Momentum", "#d62728"),
        ("p_spread_accel", "Señal 5: Spread Acceleration (Δ²RP)", "#9467bd"),
    ]
    for ax, (col, title, color) in zip(axes[:5], sig_specs):
        ax.plot(signals_df.index, signals_df[col], color=color, linewidth=1.1)
        ax.axhline(0.5, color="grey", linestyle=":", alpha=0.6)
        ax.set_ylabel("p(Risk-On)", fontsize=9)
        ax.set_title(title, loc="left", fontsize=10, fontweight="bold")
        ax.set_ylim(-0.05, 1.05)
        ax.grid(True, alpha=0.3)
    ax = axes[5]
    ax.plot(decision_df.index, decision_df["p_ensemble"], color="black", linewidth=1.5,
            label="Ensemble")
    ax.axhline(cfg.theta_in, color="green", linestyle="--", alpha=0.7,
               label=f"θ_in={cfg.theta_in}")
    ax.axhline(cfg.theta_out, color="red", linestyle="--", alpha=0.7,
               label=f"θ_out={cfg.theta_out}")
    # Sombreado de períodos Risk-On
    in_on = decision_df["regime"].values == 1
    starts, ends = [], []
    prev = False
    for i, v in enumerate(in_on):
        if v and not prev:
            starts.append(decision_df.index[i])
        if not v and prev:
            ends.append(decision_df.index[i])
        prev = v
    if prev:
        ends.append(decision_df.index[-1])
    for s, e in zip(starts, ends):
        ax.axvspan(s, e, alpha=0.12, color="green")
    ax.set_ylim(-0.05, 1.05)
    ax.set_ylabel("p_ensemble", fontsize=9)
    ax.set_title("Ensemble (verde sombreado = Risk-On)", loc="left", fontsize=10, fontweight="bold")
    ax.legend(loc="lower left", fontsize=8)
    ax.grid(True, alpha=0.3)
    ax.xaxis.set_major_locator(mdates.YearLocator(2))
    ax.xaxis.set_major_formatter(mdates.DateFormatter("%Y"))
    plt.tight_layout()
    p = f"{out}/03_signals_panel.png"
    plt.savefig(p, dpi=cfg.plot_dpi, bbox_inches="tight"); plt.close()
    plot_paths["signals"] = p

    # ----- 4. Retornos anuales -----
    yearly = summary["yearly_returns"]
    fig, ax = plt.subplots(figsize=(13, 5))
    x = np.arange(len(yearly.index))
    w = 0.4
    ax.bar(x - w/2, yearly["ret_bh"] * 100, w, color="#C00000", alpha=0.7, label="Buy & Hold")
    ax.bar(x + w/2, yearly["ret_strat"] * 100, w, color="#1F4E79", alpha=0.85, label="Estrategia")
    ax.set_xticks(x); ax.set_xticklabels(yearly.index, rotation=45)
    ax.axhline(0, color="black", linewidth=0.6)
    ax.set_ylabel("Retorno anual (%)")
    ax.set_title("Retornos anuales comparados", fontsize=13, fontweight="bold")
    ax.legend(fontsize=10)
    ax.grid(True, alpha=0.3, axis="y")
    plt.tight_layout()
    p = f"{out}/04_yearly_returns.png"
    plt.savefig(p, dpi=cfg.plot_dpi, bbox_inches="tight"); plt.close()
    plot_paths["yearly"] = p

    # ----- 5. Confusion matrix -----
    cm = summary["confusion_matrix"]
    cls = summary["classification"]
    fig, ax = plt.subplots(figsize=(7, 6))
    total = cm.sum()
    annot = np.array([[f"{cm[i,j]}\n({cm[i,j]/total:.1%})" for j in range(cm.shape[1])]
                       for i in range(cm.shape[0])])
    sns.heatmap(cm, annot=annot, fmt="", cmap="Greens", cbar=False, ax=ax,
                xticklabels=["Pred Risk-Off", "Pred Risk-On"],
                yticklabels=["True Risk-Off", "True Risk-On"],
                annot_kws={"fontsize": 13, "fontweight": "bold"})
    ax.set_title(f"Matriz de Confusión\n"
                 f"Acc={cls['Accuracy']:.1%}, F1_On={cls['F1_RiskOn']:.3f}, "
                 f"F1_Off={cls['F1_RiskOff']:.3f}, Kappa={cls['Kappa']:.3f}",
                 fontsize=11, fontweight="bold")
    plt.tight_layout()
    p = f"{out}/05_confusion_matrix.png"
    plt.savefig(p, dpi=cfg.plot_dpi, bbox_inches="tight"); plt.close()
    plot_paths["confusion"] = p

    # ----- 6. Exposure timeline -----
    fig, ax = plt.subplots(figsize=(13, 3.5))
    ax.fill_between(bt.index, 0, bt["exposure"], color="#1F4E79", alpha=0.5,
                    step="post", label="Exposición")
    ax.set_ylim(-0.05, 1.15)
    ax.set_ylabel("Exposición")
    ax.set_title("Línea temporal de exposición al Merval USD", fontsize=12, fontweight="bold")
    ax.legend(loc="upper left", fontsize=10)
    ax.grid(True, alpha=0.3)
    ax.xaxis.set_major_locator(mdates.YearLocator(2))
    ax.xaxis.set_major_formatter(mdates.DateFormatter("%Y"))
    plt.tight_layout()
    p = f"{out}/06_exposure_timeline.png"
    plt.savefig(p, dpi=cfg.plot_dpi, bbox_inches="tight"); plt.close()
    plot_paths["exposure"] = p

    # ----- 7. Rolling Sharpe (24m) -----
    win = 24
    rs_strat = bt["ret_strat"].rolling(win).mean() / bt["ret_strat"].rolling(win).std() * np.sqrt(12)
    rs_bh = bt["ret_bh"].rolling(win).mean() / bt["ret_bh"].rolling(win).std() * np.sqrt(12)
    fig, ax = plt.subplots(figsize=(13, 5))
    ax.plot(bt.index, rs_strat, color="#1F4E79", linewidth=1.8, label="Estrategia (24m)")
    ax.plot(bt.index, rs_bh, color="#C00000", linewidth=1.4, linestyle="--", label="Buy & Hold (24m)")
    ax.axhline(0, color="black", linewidth=0.6, alpha=0.5)
    ax.set_ylabel("Rolling Sharpe (24m)")
    ax.set_title("Sharpe Ratio Rolling (24 meses)", fontsize=13, fontweight="bold")
    ax.legend(loc="upper left", fontsize=10)
    ax.grid(True, alpha=0.3)
    ax.xaxis.set_major_locator(mdates.YearLocator(2))
    ax.xaxis.set_major_formatter(mdates.DateFormatter("%Y"))
    plt.tight_layout()
    p = f"{out}/07_rolling_sharpe.png"
    plt.savefig(p, dpi=cfg.plot_dpi, bbox_inches="tight"); plt.close()
    plot_paths["rolling_sharpe"] = p

    # ----- 8. Correlación entre señales -----
    fig, ax = plt.subplots(figsize=(8, 6))
    corr = signals_df.dropna().corr()
    sns.heatmap(corr, annot=True, fmt=".2f", cmap="RdBu_r", center=0, vmin=-1, vmax=1,
                ax=ax, square=True, cbar_kws={"shrink": 0.7})
    ax.set_title("Correlación entre las 5 señales", fontsize=12, fontweight="bold")
    plt.tight_layout()
    p = f"{out}/08_signal_correlation.png"
    plt.savefig(p, dpi=cfg.plot_dpi, bbox_inches="tight"); plt.close()
    plot_paths["correlation"] = p

    logger.info(f"Gráficos generados: {len(plot_paths)}")
    for name, path in plot_paths.items():
        logger.info(f"  {name:18s} → {path}")

    return plot_paths


# =============================================================================
# 8. EXPORT A EXCEL
# =============================================================================

# Styles
HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
HEADER_FONT = Font(name="Arial", size=11, bold=True, color="FFFFFF")
TITLE_FONT = Font(name="Arial", size=14, bold=True, color="1F4E79")
SUBTITLE_FONT = Font(name="Arial", size=11, bold=True, color="2E75B6")
BASE_FONT = Font(name="Arial", size=10)
POSITIVE_FONT = Font(name="Arial", size=10, color="006100")
NEGATIVE_FONT = Font(name="Arial", size=10, color="9C0006")

THIN = Side(border_style="thin", color="BFBFBF")
BOX_BORDER = Border(top=THIN, bottom=THIN, left=THIN, right=THIN)

CENTER = Alignment(horizontal="center", vertical="center")
LEFT = Alignment(horizontal="left", vertical="center")
RIGHT = Alignment(horizontal="right", vertical="center")


def _style_header_row(ws, row: int, col_start: int, col_end: int):
    for c in range(col_start, col_end + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = CENTER
        cell.border = BOX_BORDER


def _apply_box(ws, row_start: int, row_end: int, col_start: int, col_end: int):
    for r in range(row_start, row_end + 1):
        for c in range(col_start, col_end + 1):
            ws.cell(row=r, column=c).border = BOX_BORDER


def _zebra(ws, row_start: int, row_end: int, col_start: int, col_end: int):
    fill = PatternFill("solid", fgColor="F2F2F2")
    for r in range(row_start, row_end + 1):
        if (r - row_start) % 2 == 1:
            for c in range(col_start, col_end + 1):
                ws.cell(row=r, column=c).fill = fill


def _auto_width(ws):
    for col_cells in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col_cells[0].column)
        for cell in col_cells:
            if cell.value is not None:
                v = str(cell.value)
                if len(v) > max_len:
                    max_len = len(v)
        ws.column_dimensions[col_letter].width = min(max_len + 2, 35)


def export_excel(bt: pd.DataFrame, decision_df: pd.DataFrame,
                  signals_df: pd.DataFrame, summary: dict,
                  switches_log: pd.DataFrame, cfg: Config,
                  logger: logging.Logger) -> str:
    """Exporta backtest completo a Excel multi-sheet con formato profesional."""
    log_banner(logger, "PASO 9: EXPORTACIÓN A EXCEL")

    out_path = os.path.join(cfg.output_dir, cfg.excel_output)
    wb = Workbook()

    # ---------- Hoja 1: Summary ----------
    ws = wb.active
    ws.title = "Summary"
    ws["A1"] = "Modelo de Régimen Risk-On / Risk-Off - Merval USD"
    ws["A1"].font = TITLE_FONT
    ws.merge_cells("A1:D1")
    ws["A2"] = f"Generado: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
    ws["A2"].font = Font(name="Arial", size=9, italic=True, color="595959")
    ws.merge_cells("A2:D2")

    row = 4
    ws.cell(row=row, column=1, value="MÉTRICA").font = HEADER_FONT
    ws.cell(row=row, column=2, value="Estrategia").font = HEADER_FONT
    ws.cell(row=row, column=3, value="Buy & Hold").font = HEADER_FONT
    ws.cell(row=row, column=4, value="Diferencial").font = HEADER_FONT
    _style_header_row(ws, row, 1, 4)
    row += 1

    metric_rows = [
        ("CAGR", summary["strategy"]["CAGR"], summary["buy_hold"]["CAGR"], "pct"),
        ("Volatilidad anualizada", summary["strategy"]["Vol"], summary["buy_hold"]["Vol"], "pct"),
        ("Sharpe Ratio", summary["strategy"]["Sharpe"], summary["buy_hold"]["Sharpe"], "num"),
        ("Sortino Ratio", summary["strategy"]["Sortino"], summary["buy_hold"]["Sortino"], "num"),
        ("Maximum Drawdown", summary["strategy"]["MaxDD"], summary["buy_hold"]["MaxDD"], "pct"),
        ("Calmar Ratio", summary["strategy"]["Calmar"], summary["buy_hold"]["Calmar"], "num"),
        ("Skewness", summary["strategy"]["Skew"], summary["buy_hold"]["Skew"], "num"),
        ("Excess Kurtosis", summary["strategy"]["ExKurt"], summary["buy_hold"]["ExKurt"], "num"),
        ("Observaciones", summary["strategy"]["Obs"], summary["buy_hold"]["Obs"], "int"),
    ]
    for label, strat_v, bh_v, fmt in metric_rows:
        ws.cell(row=row, column=1, value=label).font = BASE_FONT
        for col, val in [(2, strat_v), (3, bh_v)]:
            c = ws.cell(row=row, column=col, value=val if pd.notna(val) else None)
            c.font = BASE_FONT
            c.alignment = RIGHT
            if fmt == "pct":
                c.number_format = "0.00%;[Red](0.00%);-"
            elif fmt == "num":
                c.number_format = "0.000;[Red](0.000);-"
            elif fmt == "int":
                c.number_format = "#,##0"
        if fmt == "pct" or fmt == "num":
            if pd.notna(strat_v) and pd.notna(bh_v):
                diff = strat_v - bh_v
                d = ws.cell(row=row, column=4, value=diff)
                d.font = POSITIVE_FONT if diff >= 0 else NEGATIVE_FONT
                d.alignment = RIGHT
                d.number_format = "+0.00%;-0.00%;-" if fmt == "pct" else "+0.000;-0.000;-"
        row += 1

    # Métricas adicionales
    row += 1
    ws.cell(row=row, column=1, value="MÉTRICAS ADICIONALES").font = SUBTITLE_FONT
    row += 1
    extras = [
        ("Information Ratio vs B&H", summary["information_ratio"], "num"),
        ("Up Capture", summary["capture"]["UpCapture"], "pct"),
        ("Down Capture", summary["capture"]["DownCapture"], "pct"),
        ("Ratio UC/DC", summary["capture"]["UC_DC_ratio"], "num"),
        ("Win Rate (monthly)", summary["capture"]["WinRate"], "pct"),
        ("# Switches", summary["n_switches"], "int"),
        ("% Tiempo en Risk-On", summary["pct_risk_on"], "pct"),
    ]
    for label, val, fmt in extras:
        ws.cell(row=row, column=1, value=label).font = BASE_FONT
        c = ws.cell(row=row, column=2, value=val if pd.notna(val) else None)
        c.font = BASE_FONT; c.alignment = RIGHT
        if fmt == "pct":
            c.number_format = "0.00%;-0.00%;-"
        elif fmt == "num":
            c.number_format = "0.000;-0.000;-"
        else:
            c.number_format = "#,##0"
        row += 1

    row += 1
    ws.cell(row=row, column=1, value="MÉTRICAS DE CLASIFICACIÓN vs GROUND TRUTH").font = SUBTITLE_FONT
    row += 1
    cls = summary["classification"]
    for k in ["Accuracy", "Kappa", "AUC_ROC",
              "F1_RiskOn", "Precision_RiskOn", "Recall_RiskOn",
              "F1_RiskOff", "Precision_RiskOff", "Recall_RiskOff"]:
        ws.cell(row=row, column=1, value=k).font = BASE_FONT
        c = ws.cell(row=row, column=2, value=cls.get(k))
        c.font = BASE_FONT; c.alignment = RIGHT
        c.number_format = "0.000"
        row += 1
    _auto_width(ws)

    # ---------- Hoja 2: Config ----------
    ws = wb.create_sheet("Config")
    ws["A1"] = "Hiperparámetros del modelo"
    ws["A1"].font = TITLE_FONT
    ws.merge_cells("A1:B1")
    row = 3
    ws.cell(row=row, column=1, value="Parámetro").font = HEADER_FONT
    ws.cell(row=row, column=2, value="Valor").font = HEADER_FONT
    _style_header_row(ws, row, 1, 2)
    row += 1
    for k, v in cfg.to_dict().items():
        ws.cell(row=row, column=1, value=k).font = BASE_FONT
        ws.cell(row=row, column=2, value=str(v)).font = BASE_FONT
        ws.cell(row=row, column=2).alignment = LEFT
        row += 1
    _zebra(ws, 4, row - 1, 1, 2)
    _apply_box(ws, 3, row - 1, 1, 2)
    _auto_width(ws)

    # ---------- Hoja 3: Signals ----------
    ws = wb.create_sheet("Signals")
    ws["A1"] = "Las 5 señales del ensemble (probabilidades de Risk-On)"
    ws["A1"].font = TITLE_FONT
    ws.merge_cells("A1:F1")
    sig_df = signals_df.copy()
    sig_df.index.name = "Fecha"
    sig_df = sig_df.reset_index()
    # Headers
    for j, col in enumerate(sig_df.columns, start=1):
        ws.cell(row=3, column=j, value=col)
    _style_header_row(ws, 3, 1, len(sig_df.columns))
    # Data
    for i, r in enumerate(sig_df.itertuples(index=False), start=4):
        for j, val in enumerate(r, start=1):
            c = ws.cell(row=i, column=j, value=val)
            c.font = BASE_FONT
            if j == 1:
                c.number_format = "yyyy-mm-dd"
            else:
                c.number_format = "0.0000"
                c.alignment = RIGHT
    # Color scale conditional formatting
    last_row = 3 + len(sig_df)
    for col_letter in ["B", "C", "D", "E", "F"]:
        rng = f"{col_letter}4:{col_letter}{last_row}"
        rule = ColorScaleRule(
            start_type="num", start_value=0, start_color="F8696B",
            mid_type="num", mid_value=0.5, mid_color="FFEB84",
            end_type="num", end_value=1, end_color="63BE7B"
        )
        ws.conditional_formatting.add(rng, rule)
    _auto_width(ws)

    # ---------- Hoja 4: Ensemble ----------
    ws = wb.create_sheet("Ensemble")
    ws["A1"] = "Probabilidad agregada y decisión binaria del ensemble"
    ws["A1"].font = TITLE_FONT
    ws.merge_cells("A1:D1")
    ens = decision_df.copy().reset_index().rename(columns={ens.columns[0] if hasattr(ens, "columns") else "index": "Fecha"}) if False else decision_df.copy()
    ens.index.name = "Fecha"
    ens = ens.reset_index()
    for j, col in enumerate(ens.columns, start=1):
        ws.cell(row=3, column=j, value=col)
    _style_header_row(ws, 3, 1, len(ens.columns))
    for i, r in enumerate(ens.itertuples(index=False), start=4):
        for j, val in enumerate(r, start=1):
            c = ws.cell(row=i, column=j, value=val)
            c.font = BASE_FONT
            if j == 1:
                c.number_format = "yyyy-mm-dd"
            elif isinstance(val, (int, float)):
                c.number_format = "0.0000"
                c.alignment = RIGHT
    _auto_width(ws)

    # ---------- Hoja 5: Backtest (mes a mes) ----------
    ws = wb.create_sheet("Backtest")
    ws["A1"] = "Backtest detallado mes a mes"
    ws["A1"].font = TITLE_FONT
    ws.merge_cells("A1:J1")
    bt_view = bt[[
        "regime_label", "exposure", "ret_merval", "ret_strat", "ret_bh",
        "tc_drag", "equity_strat", "equity_bh", "dd_strat", "dd_bh"
    ]].copy()
    bt_view.index.name = "Fecha"
    bt_view = bt_view.reset_index()
    headers = ["Fecha", "Régimen", "Exposición", "Ret. Merval", "Ret. Estrategia",
               "Ret. Buy & Hold", "Costo TC", "Equity Estrategia", "Equity B&H",
               "DD Estrategia", "DD B&H"]
    for j, h in enumerate(headers, start=1):
        ws.cell(row=3, column=j, value=h)
    _style_header_row(ws, 3, 1, len(headers))
    for i, r in enumerate(bt_view.itertuples(index=False), start=4):
        date_v, regime_v, exp_v, retm, rets, retbh, tc, eqs, eqb, dds, ddb = r
        ws.cell(row=i, column=1, value=date_v).number_format = "yyyy-mm-dd"
        ws.cell(row=i, column=2, value=regime_v).alignment = CENTER
        ws.cell(row=i, column=3, value=exp_v).number_format = "0.0"
        ws.cell(row=i, column=4, value=retm).number_format = "0.00%;[Red](0.00%);-"
        ws.cell(row=i, column=5, value=rets).number_format = "0.00%;[Red](0.00%);-"
        ws.cell(row=i, column=6, value=retbh).number_format = "0.00%;[Red](0.00%);-"
        ws.cell(row=i, column=7, value=tc).number_format = "0.000%"
        ws.cell(row=i, column=8, value=eqs).number_format = "$#,##0.00"
        ws.cell(row=i, column=9, value=eqb).number_format = "$#,##0.00"
        ws.cell(row=i, column=10, value=dds).number_format = "0.00%;[Red](0.00%);-"
        ws.cell(row=i, column=11, value=ddb).number_format = "0.00%;[Red](0.00%);-"
        for j in range(1, 12):
            ws.cell(row=i, column=j).font = BASE_FONT
        # Color cell de régimen
        regime_cell = ws.cell(row=i, column=2)
        if regime_v == "Risk-On":
            regime_cell.fill = PatternFill("solid", fgColor="C6EFCE")
            regime_cell.font = Font(name="Arial", size=10, bold=True, color="006100")
        else:
            regime_cell.fill = PatternFill("solid", fgColor="FFC7CE")
            regime_cell.font = Font(name="Arial", size=10, bold=True, color="9C0006")
    ws.freeze_panes = "B4"
    _auto_width(ws)

    # ---------- Hoja 6: Switches ----------
    ws = wb.create_sheet("Switches")
    ws["A1"] = "Log de switches del modelo"
    ws["A1"].font = TITLE_FONT
    ws.merge_cells("A1:E1")
    if len(switches_log) > 0:
        sw = switches_log.copy()
        headers = list(sw.columns)
        for j, h in enumerate(headers, start=1):
            ws.cell(row=3, column=j, value=h)
        _style_header_row(ws, 3, 1, len(headers))
        for i, r in enumerate(sw.itertuples(index=False), start=4):
            for j, val in enumerate(r, start=1):
                c = ws.cell(row=i, column=j, value=val)
                c.font = BASE_FONT
                if isinstance(val, (pd.Timestamp, datetime)):
                    c.number_format = "yyyy-mm-dd"
                elif isinstance(val, float):
                    c.number_format = "0.0000"
        _zebra(ws, 4, 3 + len(sw), 1, len(headers))
    else:
        ws.cell(row=3, column=1, value="(Sin switches en el período)").font = BASE_FONT
    _auto_width(ws)

    # ---------- Hoja 7: Yearly Returns ----------
    ws = wb.create_sheet("Yearly Returns")
    ws["A1"] = "Retornos anuales: Estrategia vs Buy & Hold"
    ws["A1"].font = TITLE_FONT
    ws.merge_cells("A1:E1")
    yr = summary["yearly_returns"].reset_index()
    headers = ["Año", "Ret. Estrategia", "Ret. Buy & Hold", "Alpha", "% Risk-On", "# Switches"]
    yr = yr.rename(columns={"year": "Año"})
    yr_cols = ["Año", "ret_strat", "ret_bh", "alpha", "pct_on", "n_switches"]
    for j, h in enumerate(headers, start=1):
        ws.cell(row=3, column=j, value=h)
    _style_header_row(ws, 3, 1, len(headers))
    for i, r in enumerate(yr[yr_cols].itertuples(index=False), start=4):
        year, rs, rb, al, pct, sw = r
        ws.cell(row=i, column=1, value=str(int(year))).alignment = CENTER
        ws.cell(row=i, column=2, value=rs).number_format = "0.00%;[Red](0.00%);-"
        ws.cell(row=i, column=3, value=rb).number_format = "0.00%;[Red](0.00%);-"
        al_c = ws.cell(row=i, column=4, value=al)
        al_c.number_format = "+0.00%;-0.00%;-"
        al_c.font = POSITIVE_FONT if al >= 0 else NEGATIVE_FONT
        ws.cell(row=i, column=5, value=pct).number_format = "0.0%"
        ws.cell(row=i, column=6, value=sw).alignment = CENTER
        for j in range(1, 7):
            if j not in (4,):
                ws.cell(row=i, column=j).font = BASE_FONT
    _zebra(ws, 4, 3 + len(yr), 1, 6)
    _apply_box(ws, 3, 3 + len(yr), 1, 6)
    _auto_width(ws)

    # ---------- Hoja 8: Top Drawdowns ----------
    ws = wb.create_sheet("Drawdowns")
    ws["A1"] = "Top 10 drawdowns de la estrategia"
    ws["A1"].font = TITLE_FONT
    ws.merge_cells("A1:F1")
    td = summary["top_drawdowns"]
    if len(td) > 0:
        headers = list(td.columns)
        for j, h in enumerate(headers, start=1):
            ws.cell(row=3, column=j, value=h)
        _style_header_row(ws, 3, 1, len(headers))
        for i, r in enumerate(td.itertuples(index=False), start=4):
            for j, val in enumerate(r, start=1):
                c = ws.cell(row=i, column=j, value=val)
                c.font = BASE_FONT
                if isinstance(val, float) and j == 4:  # max_dd column
                    c.number_format = "0.00%"
                    c.font = NEGATIVE_FONT
                elif isinstance(val, float):
                    c.number_format = "0.0"
        _zebra(ws, 4, 3 + len(td), 1, len(headers))
        _apply_box(ws, 3, 3 + len(td), 1, len(headers))
    _auto_width(ws)

    # ---------- Hoja 9: Confusion Matrix ----------
    ws = wb.create_sheet("Confusion Matrix")
    ws["A1"] = "Matriz de confusión vs Ground Truth (Triple Barrier ex-post)"
    ws["A1"].font = TITLE_FONT
    ws.merge_cells("A1:D1")
    cm = summary["confusion_matrix"]
    total = cm.sum()
    ws["B3"] = "Pred Risk-Off"
    ws["C3"] = "Pred Risk-On"
    ws["A4"] = "True Risk-Off"
    ws["A5"] = "True Risk-On"
    for cell in [ws["B3"], ws["C3"]]:
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = CENTER
    for cell in [ws["A4"], ws["A5"]]:
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = CENTER
    ws["B4"] = int(cm[0,0]); ws["C4"] = int(cm[0,1])
    ws["B5"] = int(cm[1,0]); ws["C5"] = int(cm[1,1])
    for r in [4, 5]:
        for c in ["B", "C"]:
            cell = ws[f"{c}{r}"]
            cell.alignment = CENTER
            cell.font = Font(name="Arial", size=12, bold=True)
            cell.border = BOX_BORDER
    # Heat fill for diagonal
    ws["B4"].fill = PatternFill("solid", fgColor="C6EFCE")
    ws["C5"].fill = PatternFill("solid", fgColor="C6EFCE")
    ws["B5"].fill = PatternFill("solid", fgColor="FFC7CE")
    ws["C4"].fill = PatternFill("solid", fgColor="FFC7CE")

    # Métricas debajo
    row = 8
    ws.cell(row=row, column=1, value="MÉTRICAS DERIVADAS").font = SUBTITLE_FONT
    row += 1
    for k in ["Accuracy", "Kappa", "AUC_ROC",
              "F1_RiskOn", "Precision_RiskOn", "Recall_RiskOn",
              "F1_RiskOff", "Precision_RiskOff", "Recall_RiskOff"]:
        ws.cell(row=row, column=1, value=k).font = BASE_FONT
        c = ws.cell(row=row, column=2, value=cls.get(k))
        c.font = BASE_FONT; c.alignment = RIGHT; c.number_format = "0.000"
        row += 1
    _auto_width(ws)

    # ---------- Hoja 10: Statistics ----------
    ws = wb.create_sheet("Statistics")
    ws["A1"] = "Estadísticas descriptivas de los inputs y outputs"
    ws["A1"].font = TITLE_FONT
    ws.merge_cells("A1:H1")
    desc = pd.concat([signals_df, decision_df[["p_ensemble", "regime"]]], axis=1).describe()
    desc.index.name = "Statistic"
    desc = desc.reset_index()
    for j, col in enumerate(desc.columns, start=1):
        ws.cell(row=3, column=j, value=col)
    _style_header_row(ws, 3, 1, len(desc.columns))
    for i, r in enumerate(desc.itertuples(index=False), start=4):
        for j, val in enumerate(r, start=1):
            c = ws.cell(row=i, column=j, value=val)
            c.font = BASE_FONT
            if isinstance(val, float):
                c.number_format = "0.0000"
                c.alignment = RIGHT
    _zebra(ws, 4, 3 + len(desc), 1, len(desc.columns))
    _apply_box(ws, 3, 3 + len(desc), 1, len(desc.columns))
    _auto_width(ws)

    # Save
    wb.save(out_path)
    logger.info(f"Excel guardado: {out_path}")
    logger.info(f"  Hojas: {wb.sheetnames}")
    return out_path


# =============================================================================
# 9. MAIN PIPELINE
# =============================================================================

def run_production(cfg: Config = CONFIG):
    """Ejecuta el pipeline completo de producción."""
    os.makedirs(cfg.output_dir, exist_ok=True)
    log_path = os.path.join(cfg.output_dir, cfg.log_file)
    logger = setup_logger(log_path)

    log_banner(logger, "MODELO DE RÉGIMEN RISK-ON / RISK-OFF MERVAL USD", char="=", width=80)
    logger.info(f"Inicio de la ejecución: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    logger.info(f"Output directory: {cfg.output_dir}")
    logger.info(f"Log file: {log_path}")

    # PASO 1: Inputs
    df = load_inputs(cfg, logger)
    rets = df["ret"]

    # PASO 2-3: MS con walk-forward refit
    p_ms = walkforward_ms_refit(rets, cfg, logger)

    # PASO 4: Señales
    signals_df = build_all_signals(df, cfg, logger, p_ms)

    # PASO 5: Ensemble
    p_ensemble = build_ensemble(signals_df, logger)

    # PASO 6: Decisión
    decision_df, switches_log = apply_hysteresis(p_ensemble, cfg, logger)

    # PASO 7: Backtest
    bt = run_backtest(decision_df, rets, cfg, logger)

    # PASO 8: Ground truth + métricas
    logger.info("\nGenerando ground truth con Triple Barrier...")
    y_true = triple_barrier_labels(df["merval"], cfg)
    summary = consolidate_metrics(bt, decision_df, y_true, logger)

    # PASO 9: Gráficos
    plot_paths = generate_plots(bt, decision_df, signals_df, summary, cfg, logger)

    # PASO 10: Excel
    excel_path = export_excel(bt, decision_df, signals_df, summary,
                                switches_log, cfg, logger)

    # Cierre
    log_banner(logger, "EJECUCIÓN COMPLETADA", char="=", width=80)
    logger.info(f"\nResumen final:")
    logger.info(f"  CAGR Estrategia: {summary['strategy']['CAGR']:.2%}")
    logger.info(f"  CAGR Buy & Hold: {summary['buy_hold']['CAGR']:.2%}")
    logger.info(f"  Sharpe Estrategia: {summary['strategy']['Sharpe']:.3f}")
    logger.info(f"  Information Ratio: {summary['information_ratio']:.3f}")
    logger.info(f"  MaxDD Estrategia: {summary['strategy']['MaxDD']:.2%}")
    logger.info(f"  # Switches: {summary['n_switches']}")
    logger.info(f"\nArchivos generados en: {cfg.output_dir}")

    return {
        "config": cfg, "summary": summary,
        "backtest": bt, "decision": decision_df, "signals": signals_df,
        "switches": switches_log, "excel_path": excel_path,
        "plot_paths": plot_paths,
    }


if __name__ == "__main__":
    results = run_production()
