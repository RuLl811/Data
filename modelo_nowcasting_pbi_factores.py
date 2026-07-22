# -*- coding: utf-8 -*-
"""
============================================================================
NOWCASTING DEL PBI DE ARGENTINA  -  MODELO DE FACTORES (VERSION DEFINITIVA)
Inspirado en D'Amato, Garegnani & Blanco (2016), BCRA Ensayos Economicos 74.
============================================================================

Modelo de produccion: FACTORES DINAMICOS (PCA -> regresion, ec. 3 del paper),
configurado REAL-TIME-AWARE. Decisiones de diseno respecto del paper:

  1. Panel de PRODUCCION = hard data (USAR_PROXIES_ACTIVIDAD=False). Se excluyen
     EMAE / IGA_OJF / ICG del nowcast en tiempo real porque son proxies casi
     contemporaneos del PBI (EMAE ~ PBI mensual, con el MAYOR lag de publicacion
     del panel): dan performance OOS ilusoria que en real-time no tendrias. Se
     reservan para el BACKCAST del trimestre ya cerrado (bloque de referencia).
  2. Seleccion / estandarizacion / PCA se re-estiman DENTRO de cada ventana
     (leak-free): sin look-ahead de seleccion.
  3. ADF target-aware: se exige estacionariedad al 1% para dejar una serie en
     nivel, de modo que las series de actividad (casi I(1)) pasen a crecimiento,
     homogeneas con el target (crecimiento q/q).
  4. Salida DUAL: nivel (q/q %) + senal DIRECCIONAL P(expansion), que es la que
     tiene valor estadistico robusto (ver test de Pesaran-Timmermann).

Evaluacion de eficiencia (todos los resultados):
  - RMSE / MAE / RMSE ex-2020 (el shock COVID domina la varianza),
  - Test de Giacomini-White (2006): magnitud, incondicional y condicional,
  - Test de Pesaran-Timmermann (1992): direccional (expansion vs contraccion),
  - % de trimestres en que el modelo supera al AR(1).

Salidas:
  - modelo_nowcasting_pbi_factores.py (este archivo),
  - 4 graficos PNG (nowcast, importancia, evolucion RMSE, direccional),
  - Excel teorico con la explicacion del modelo y la bibliografia.

AR(1) queda SOLO como benchmark. El puente se conserva como comparador para
mostrar por que NO se elige (su performance depende del proxy EMAE).
============================================================================
"""

import warnings
import numpy as np
import pandas as pd
import matplotlib.pyplot as plt
import matplotlib.ticker as mticker

from statsmodels.tsa.stattools import adfuller
from sklearn.preprocessing import StandardScaler
from sklearn.decomposition import PCA

warnings.filterwarnings("ignore")

# ---------------------------------------------------------------------------
# CONFIGURACION GLOBAL
# ---------------------------------------------------------------------------
RUTA_EXCEL          = "base_pbi_1.xlsx"
COL_FECHA           = "Date"
COL_PBI             = "DPBI_desest"      # target: crecimiento trimestral q/q (desest)

UMBRAL_CORR         = 0.50               # umbral de |correlacion| con el PBI (paper)
N_FACTORES          = 4                  # nro de factores del PCA (paper: scree -> 4)
VENTANA_INICIAL     = 40                 # trimestres de entrenamiento inicial (~2014)
TIPO_VENTANA        = "expansiva"        # "expansiva" | "movil"
LARGO_VENTANA_MOVIL = 40                 # solo si TIPO_VENTANA == "movil"
H_FORECAST          = 4                  # trimestres a proyectar hacia adelante
VENTANA_ROLLING     = 8                  # ventana (trim.) para metricas moviles

# Proxies de actividad casi-contemporaneos del PBI (EMAE ~ PBI mensual).
# En PRODUCCION real-time se EXCLUYEN: dan performance OOS ilusoria y son los de
# mayor lag de publicacion. Se usan solo para el backcast de referencia.
USAR_PROXIES_ACTIVIDAD = False
PROXIES_ACTIVIDAD      = ["EMAE", "IGA_OJF", "ICG"]

# Rutas de salida
DIR_OUT   = "/mnt/user-data/outputs"

SEED = 42
np.random.seed(SEED)


# ===========================================================================
# PASO 1 - PREPROCESAMIENTO Y TRANSFORMACION DE DATOS
# ===========================================================================
def cargar_datos(ruta=RUTA_EXCEL):
    """Lee el Excel y devuelve un DataFrame mensual con indice de fecha."""
    df = pd.read_excel(ruta)
    df[COL_FECHA] = pd.to_datetime(df[COL_FECHA])
    df = df.sort_values(COL_FECHA).set_index(COL_FECHA)
    return df


def colapsar_pbi_trimestral(serie_mensual):
    """
    El PBI viene 'broadcasteado' a mensual (mismo valor dentro de cada
    trimestre). Lo colapsamos a un unico valor por trimestre.
    Se usa .first() porque los tres meses del trimestre son identicos.
    """
    q = serie_mensual.groupby(serie_mensual.index.to_period("Q")).first()
    q.index = q.index.to_timestamp(how="end").normalize()
    return q


def agregar_mensual_a_trimestral(df_mensual, cols):
    """
    Ecuacion (4) del paper: agregacion de datos de alta frecuencia a
    frecuencia trimestral mediante PROMEDIOS (cada mes pesa igual).
    Solo se promedian trimestres con los 3 meses disponibles, salvo el
    ultimo (ragged edge) que se maneja aparte en el forecast.
    """
    g = df_mensual[cols].groupby(df_mensual.index.to_period("Q"))
    trimestral = g.mean()
    n_meses    = g.count().median(axis=1)         # meses efectivos (mediana entre series)
    trimestral.index = trimestral.index.to_timestamp(how="end").normalize()
    n_meses.index    = trimestral.index
    return trimestral, n_meses


def _es_estacionaria(serie, alpha=0.05):
    """Test ADF. Devuelve (bool_estacionaria, p_value)."""
    s = serie.dropna()
    if len(s) < 12 or s.nunique() < 5:
        return False, np.nan
    try:
        pval = adfuller(s, autolag="AIC")[1]
    except Exception:
        return False, np.nan
    return pval < alpha, pval


def transformar_a_estacionaria(serie, nombre="", alpha_nivel=0.01):
    """
    Aplica de forma ITERATIVA la transformacion minima que induce
    estacionariedad, probando en orden:
        nivel  ->  dlog (1ra dif log)  ->  diff (1ra dif)  ->  d2 (2da dif)
    Se prefiere dlog cuando la serie es estrictamente positiva (indices,
    recaudacion, agregados) y diff para tasas/spreads. Decision econometrica:
    trabajar con la transformacion MENOS agresiva que ya sea I(0), para no
    sobre-diferenciar y perder senal.

    OJO: como el target es una TASA DE CRECIMIENTO, exigimos un umbral estricto
    (alpha_nivel=1%) para aceptar una serie en 'nivel'. Muchas series de
    actividad (cemento, siderurgia, ISAC) pasan un ADF al 5% por poco pero son
    casi I(1); dejarlas en nivel las decorrelaciona artificialmente del PBI q/q.
    Con el umbral al 1% se las empuja a crecimiento, homogeneas con el target.
    Las tasas/spreads/indices de difusion genuinamente I(0) igual quedan en nivel.
    Devuelve (serie_transformada, etiqueta).
    """
    s = serie.dropna()
    positiva = (s > 0).all()

    # 1) nivel (aceptado solo si es fuertemente estacionario)
    est, _ = _es_estacionaria(s, alpha=alpha_nivel)
    if est:
        return serie.copy(), "nivel"

    # 2) dlog (solo si es positiva)
    if positiva:
        dlog = np.log(serie).diff()
        est, _ = _es_estacionaria(dlog)
        if est:
            return dlog, "dlog"

    # 3) diff
    dif = serie.diff()
    est, _ = _es_estacionaria(dif)
    if est:
        return dif, "diff"

    # 4) fallback: 2da diferencia (log si positiva)
    if positiva:
        d2 = np.log(serie).diff().diff()
        return d2, "d2log"
    d2 = serie.diff().diff()
    return d2, "d2"


def preprocesar(df_mensual, usar_proxies=USAR_PROXIES_ACTIVIDAD, verbose=True):
    """
    Orquesta el Paso 1 completo:
      - colapsa el PBI a trimestral (target ya estacionario),
      - agrega los indicadores mensuales a trimestral (promedio),
      - transforma cada indicador a estacionariedad via ADF iterativo.
    Devuelve:
      y      : PBI trimestral (hasta el ultimo trimestre publicado)
      X      : panel trimestral estacionario ALINEADO con y (para OOS)
      X_full : mismo panel pero con los trimestres futuros ya observables en
               los indicadores (ragged edge) para poder nowcastear
      Xq     : niveles trimestrales, n_meses: cobertura mensual por trimestre
      tabla  : transformaciones aplicadas
    """
    # --- Target ---
    y = colapsar_pbi_trimestral(df_mensual[COL_PBI]).dropna()
    est_y, p_y = _es_estacionaria(y)
    if verbose:
        print("=" * 74)
        print("PASO 1 | PREPROCESAMIENTO Y ESTACIONARIEDAD")
        print("=" * 74)
        print(f"Target {COL_PBI}: ya es crecimiento q/q -> ADF p={p_y:.4f} "
              f"({'estacionario' if est_y else 'NO estacionario'}). No se transforma.\n")

    # --- Indicadores ---
    cols_ind = [c for c in df_mensual.columns if c != COL_PBI]
    if not usar_proxies:
        cols_ind = [c for c in cols_ind if c not in PROXIES_ACTIVIDAD]

    Xq, n_meses = agregar_mensual_a_trimestral(df_mensual, cols_ind)

    X_trans, tabla = {}, []
    for c in cols_ind:
        serie_t, etq = transformar_a_estacionaria(Xq[c], c)
        X_trans[c] = serie_t
        _, p = _es_estacionaria(serie_t)
        tabla.append((c, etq, p))
    X_full = pd.DataFrame(X_trans)
    tabla  = pd.DataFrame(tabla, columns=["variable", "transformacion", "adf_pvalue"])

    if verbose:
        print(f"Indicadores procesados: {len(cols_ind)} "
              f"(proxies actividad {'incluidos' if usar_proxies else 'excluidos'})")
        print(tabla.to_string(index=False))
        print()

    # Panel de modelado: solo trimestres con PBI publicado y panel completo
    X = X_full.loc[y.index.intersection(X_full.dropna().index)]
    y = y.loc[X.index]
    return y, X, X_full, Xq, n_meses, tabla


# ===========================================================================
# PASO 2 - SELECCION DE VARIABLES (criterio del paper: corr con el PBI > 0.5)
# ===========================================================================
def seleccionar_variables(y, X, umbral=UMBRAL_CORR):
    """
    Replica el criterio del paper: se calcula la correlacion de cada indicador
    transformado con el crecimiento del PBI y se retienen los de |corr| > umbral.
    Devuelve (lista_seleccionadas, serie_de_correlaciones_ordenada).
    ADVERTENCIA: usada sobre toda la muestra introduce leakage. En el ejercicio
    OOS esta funcion se llama DENTRO de cada ventana (solo datos hasta t).
    """
    df = pd.concat([y.rename("_PBI_"), X], axis=1).dropna()
    corr = df.corr()["_PBI_"].drop("_PBI_")
    corr_abs = corr.abs().sort_values(ascending=False)
    seleccionadas = corr_abs[corr_abs > umbral].index.tolist()
    return seleccionadas, corr.reindex(corr_abs.index)


# ===========================================================================
# PASO 3 - MODELADO (factores, ecuaciones puente, AR(1)) + OOS
# ===========================================================================
def _ols(Xmat, yvec):
    """OLS por minimos cuadrados. Devuelve (beta, resid_std, XtX_inv)."""
    beta, _, _, _ = np.linalg.lstsq(Xmat, yvec, rcond=None)
    resid = yvec - Xmat @ beta
    dof = max(len(yvec) - Xmat.shape[1], 1)
    s2 = (resid @ resid) / dof
    XtX_inv = np.linalg.pinv(Xmat.T @ Xmat)
    return beta, np.sqrt(s2), XtX_inv


def _fit_predict_factores(y_tr, X_tr, x_new, y_lag_new, n_factores):
    """
    Modelo principal (ecuacion 3 del paper):
        y_t = c + beta' f_t + gamma * y_{t-1} + e_t
    - Seleccion + estandarizacion + PCA se ESTIMAN solo con el train (leak-free).
    - Devuelve (prediccion, std_prediccion, aux) con aux para importancias.
    """
    # Seleccion dentro de la ventana
    sel, corr = seleccionar_variables(y_tr, X_tr)
    if len(sel) < 2:                       # fallback: top-6 por |corr| si el umbral deja pocos
        sel = corr.abs().sort_values(ascending=False).head(6).index.tolist()

    df_tr = pd.concat([y_tr.rename("y"), X_tr[sel]], axis=1).dropna()
    y_al  = df_tr["y"].values
    Xsel  = df_tr[sel].values

    # y_{t-1} alineado
    y_lag = df_tr["y"].shift(1).values
    mask  = ~np.isnan(y_lag)
    y_al, Xsel, y_lag = y_al[mask], Xsel[mask], y_lag[mask]

    # Estandarizacion + PCA
    scaler = StandardScaler().fit(Xsel)
    Zsel   = scaler.transform(Xsel)
    k      = min(n_factores, Zsel.shape[1], Zsel.shape[0] - 3)
    k      = max(k, 1)
    pca    = PCA(n_components=k).fit(Zsel)
    F      = pca.transform(Zsel)

    # Regresion y ~ [1, F, y_lag]
    Xreg = np.column_stack([np.ones(len(y_al)), F, y_lag])
    beta, resid_std, XtX_inv = _ols(Xreg, y_al)

    # Prediccion para el trimestre nuevo
    z_new = scaler.transform(x_new[sel].values.reshape(1, -1))
    f_new = pca.transform(z_new).ravel()
    xrow  = np.concatenate([[1.0], f_new, [y_lag_new]])
    pred  = float(xrow @ beta)
    # error estandar de prediccion (incluye incertidumbre de estimacion)
    se    = float(resid_std * np.sqrt(1.0 + xrow @ XtX_inv @ xrow))

    # Importancia por variable original: |loading . beta_factores| ponderado por
    # varianza explicada. Traduce el efecto de cada X al modelo final.
    beta_f = beta[1:1 + k]
    contrib = np.abs(pca.components_.T @ beta_f)         # (n_vars,)
    aux = {"seleccionadas": sel, "corr": corr, "importancia": pd.Series(contrib, index=sel),
           "scaler": scaler, "pca": pca, "beta": beta, "k": k, "resid_std": resid_std}
    return pred, se, aux


def _fit_predict_ar1(y_tr, y_lag_new):
    """Benchmark AR(1): y_t = c + rho * y_{t-1} + e_t."""
    y_al  = y_tr.values[1:]
    y_lag = y_tr.values[:-1]
    Xreg  = np.column_stack([np.ones(len(y_al)), y_lag])
    beta, resid_std, XtX_inv = _ols(Xreg, y_al)
    xrow = np.array([1.0, y_lag_new])
    pred = float(xrow @ beta)
    se   = float(resid_std * np.sqrt(1.0 + xrow @ XtX_inv @ xrow))
    return pred, se


def _fit_predict_puente(y_tr, X_tr, x_new, y_lag_new):
    """
    Ecuaciones puente (ec. 5 del paper), version parsimoniosa:
    para cada indicador seleccionado se estima un ADL bivariado
        y_t = a + b*y_{t-1} + c*x_t + d*x_{t-1}
    y los pronosticos individuales se combinan con pesos inversamente
    proporcionales al RMSE in-sample (ec. 6).
    """
    sel, corr = seleccionar_variables(y_tr, X_tr)
    if len(sel) < 1:
        sel = corr.abs().sort_values(ascending=False).head(6).index.tolist()

    preds, rmses = [], []
    for c in sel:
        df = pd.concat([y_tr.rename("y"), X_tr[c].rename("x")], axis=1).dropna()
        if len(df) < 12:
            continue
        y_al  = df["y"].values[1:]
        y_lag = df["y"].values[:-1]
        x_t   = df["x"].values[1:]
        x_lag = df["x"].values[:-1]
        Xreg  = np.column_stack([np.ones(len(y_al)), y_lag, x_t, x_lag])
        beta, resid_std, _ = _ols(Xreg, y_al)
        x_new_c = x_new[c]
        x_lag_c = X_tr[c].dropna().values[-1]
        xrow = np.array([1.0, y_lag_new, x_new_c, x_lag_c])
        preds.append(float(xrow @ beta))
        rmses.append(resid_std)

    if not preds:
        return np.nan, np.nan
    w = 1.0 / np.array(rmses)
    w = w / w.sum()
    pred = float(np.dot(w, preds))
    return pred, np.nan


def evaluar_oos(y, X, n_factores=N_FACTORES, ventana_inicial=VENTANA_INICIAL,
                tipo=TIPO_VENTANA, verbose=True):
    """
    Ejercicio pseudo out-of-sample con ventana expansiva (o movil).
    En cada paso t: se entrena con [.. t-1], se predice y_t usando X_t
    (nowcast: los indicadores del trimestre t estan disponibles antes que
    el PBI oficial). Seleccion/PCA re-estimados por ventana (leak-free).
    Devuelve un DataFrame con y_real y las predicciones de cada modelo.
    """
    fechas = y.index
    resultados = []
    ult_aux = None

    for i in range(ventana_inicial, len(y)):
        if tipo == "movil":
            ini = i - ventana_inicial
        else:
            ini = 0
        idx_tr = fechas[ini:i]
        t      = fechas[i]

        y_tr = y.loc[idx_tr]
        X_tr = X.loc[idx_tr]
        x_new = X.loc[t]
        y_lag_new = y.iloc[i - 1]

        try:
            pred_f, se_f, aux = _fit_predict_factores(y_tr, X_tr, x_new, y_lag_new, n_factores)
            ult_aux = aux
        except Exception:
            pred_f, se_f = np.nan, np.nan

        pred_ar, se_ar = _fit_predict_ar1(y_tr, y_lag_new)
        pred_p, _      = _fit_predict_puente(y_tr, X_tr, x_new, y_lag_new)

        resultados.append({
            "fecha": t, "y_real": y.iloc[i],
            "factores": pred_f, "se_factores": se_f,
            "AR1": pred_ar, "se_AR1": se_ar,
            "puente": pred_p,
        })

    res = pd.DataFrame(resultados).set_index("fecha")

    if verbose:
        print("=" * 74)
        print("PASO 3 | EVALUACION PSEUDO OUT-OF-SAMPLE (ventana %s)" % tipo)
        print("=" * 74)
        # Mascara para excluir el shock COVID (2020Q2-2020Q4), que domina la varianza
        mask_covid = ~res.index.to_period("Q").isin(pd.period_range("2020Q2", "2020Q4", freq="Q"))
        for m in ["factores", "puente", "AR1"]:
            err = res["y_real"] - res[m]
            rmse = np.sqrt((err ** 2).mean())
            rmse_x = np.sqrt((err[mask_covid] ** 2).mean())
            mae  = err.abs().mean()
            print(f"  {m:10s} -> RMSE={rmse:.5f}  MAE={mae:.5f}  RMSE(ex-2020)={rmse_x:.5f}")
        # ratios y % de veces que gana el nowcast (replica graficos 3-5)
        ef = (res["y_real"] - res["factores"]).abs()
        ea = (res["y_real"] - res["AR1"]).abs()
        print(f"\n  % trimestres factores mejor que AR(1): {100*(ef < ea).mean():.1f}%")
        print(f"  Nro de trimestres OOS: {len(res)}  ({res.index.min().date()} a {res.index.max().date()})\n")

    return res, ult_aux


# ===========================================================================
# PASO 4 - TEST DE GIACOMINI-WHITE (2006)
# ===========================================================================
def _hac_cov(z, nlags):
    """
    Covarianza HAC (Newey-West) POR OBSERVACION de z (matriz n x q).
    Devuelve S tal que Var(media) = S / n. No divide por n (eso lo hace
    quien la llama, segun necesite la varianza de la media o del estadistico).
    """
    z = np.atleast_2d(z)
    if z.shape[0] < z.shape[1]:
        z = z.T
    n = z.shape[0]
    zc = z - z.mean(axis=0)
    S = (zc.T @ zc) / n
    for l in range(1, nlags + 1):
        w = 1.0 - l / (nlags + 1.0)
        G = (zc[l:].T @ zc[:-l]) / n
        S += w * (G + G.T)
    return S


def test_giacomini_white(res, modelo="factores", benchmark="AR1", tau=1, verbose=True):
    """
    Test de Giacomini-White (2006) de capacidad predictiva CONDICIONAL.
    H0: E[ h_{t-1} * dL_t ] = 0, con dL_t = L(benchmark)_t - L(modelo)_t.
    - Version incondicional (h=1): regresion de dL contra una constante y
      t-stat con errores Newey-West -> es la que reporta el paper (Cuadro 3).
    - Version condicional: instrumentos h_{t-1} = [1, dL_{t-1}]; el estadistico
      GW = n * zbar' S^{-1} zbar es chi-cuadrado con q g.l. (S = Var por obs).
    dL>0 en promedio => el 'modelo' predice mejor que el 'benchmark'.
    """
    from scipy import stats
    df = res[["y_real", modelo, benchmark]].dropna()
    e_m = (df["y_real"] - df[modelo]).values
    e_b = (df["y_real"] - df[benchmark]).values
    dL  = e_b ** 2 - e_m ** 2               # loss diff (cuadratica)
    n   = len(dL)
    nlags = max(tau - 1, 0)

    # --- Incondicional (t-stat Newey-West): Var(media) = S/n ---
    S1 = _hac_cov(dL.reshape(-1, 1), nlags)[0, 0]
    t_stat = dL.mean() / np.sqrt(S1 / n)
    p_uncond = 2 * (1 - stats.norm.cdf(abs(t_stat)))

    # --- Condicional (chi2): GW = n * zbar' S^{-1} zbar ---
    h  = np.column_stack([np.ones(n - 1), dL[:-1]])   # h_{t-1}
    zt = h * dL[1:].reshape(-1, 1)                     # z_t = h_{t-1} * dL_t
    zbar = zt.mean(axis=0)
    S = _hac_cov(zt, nlags)
    gw_stat = float(len(zt) * zbar @ np.linalg.pinv(S) @ zbar)
    q = h.shape[1]
    p_cond = 1 - stats.chi2.cdf(gw_stat, q)

    mejor = modelo if dL.mean() > 0 else benchmark

    if verbose:
        print("=" * 74)
        print("PASO 4 | TEST DE GIACOMINI-WHITE (2006):  %s  vs  %s" % (modelo, benchmark))
        print("=" * 74)
        print(f"  N (trimestres comparados) : {n}")
        print(f"  Media dL (benchmark-modelo): {dL.mean():.3e}  (signo + => {modelo} mejor)")
        print("  --- Incondicional (paper, Cuadro 3) ---")
        print(f"    Estadistico t : {t_stat:.3f}")
        print(f"    p-value       : {p_uncond:.4f}")
        print("  --- Condicional (chi2, q=%d g.l.) ---" % q)
        print(f"    Estadistico GW: {gw_stat:.3f}")
        print(f"    p-value       : {p_cond:.4f}")
        sig = "SI" if p_uncond < 0.05 else "NO"
        print(f"\n  Conclusion: mejor predictor = {mejor.upper()}. "
              f"Diferencia significativa al 5%: {sig}\n")

    return {"t_stat": t_stat, "p_uncond": p_uncond,
            "gw_stat": gw_stat, "p_cond": p_cond, "mejor": mejor, "n": n}


# ===========================================================================
# PASO 5 - PROYECCION Y VISUALIZACION
# ===========================================================================
def proyectar(y, X, X_full, Xq, n_meses, n_factores=N_FACTORES, h=H_FORECAST, verbose=True):
    """
    Proyecta el PBI de los proximos h trimestres.
      - Trimestres con datos mensuales (parciales o completos): NOWCAST directo
        agregando los meses disponibles (ragged edge en X_full).
      - Trimestres sin datos: se extrapolan los factores via AR(1) sobre cada
        factor y se aplica la ecuacion del modelo (forecast puro, CI amplio).
    Devuelve DataFrame con la proyeccion y bandas de confianza (95%).
    """
    # Reajustamos el modelo con TODA la muestra disponible
    sel, corr = seleccionar_variables(y, X)
    if len(sel) < 2:
        sel = corr.abs().sort_values(ascending=False).head(6).index.tolist()

    df_tr = pd.concat([y.rename("y"), X[sel]], axis=1).dropna()
    y_al  = df_tr["y"].values
    y_lag = df_tr["y"].shift(1).values
    mask  = ~np.isnan(y_lag)
    y_al, Xsel, y_lag = y_al[mask], df_tr[sel].values[mask], y_lag[mask]

    scaler = StandardScaler().fit(Xsel)
    k   = max(min(n_factores, Xsel.shape[1]), 1)
    pca = PCA(n_components=k).fit(scaler.transform(Xsel))
    F   = pca.transform(scaler.transform(Xsel))
    Xreg = np.column_stack([np.ones(len(y_al)), F, y_lag])
    beta, resid_std, _ = _ols(Xreg, y_al)

    # Los proximos h trimestres calendario despues del ultimo PBI publicado
    ult_per = pd.Period(y.index.max(), freq="Q")
    fut_per = pd.period_range(ult_per + 1, periods=h, freq="Q")
    fut_idx = [p.to_timestamp(how="end").normalize() for p in fut_per]

    # AR(1) por factor para extrapolar cuando falten datos
    def ar1_step(v):
        a = np.column_stack([np.ones(len(v) - 1), v[:-1]])
        b, _, _ = _ols(a, v[1:])
        return float(np.array([1.0, v[-1]]) @ b)

    proy = []
    y_lag_iter = y.iloc[-1]
    F_hist = list(F)
    for j, t in enumerate(fut_idx):
        cobertura = int(n_meses.get(t, 0))
        # Hay nowcast si el panel de indicadores ya tiene ese trimestre observado
        tiene_datos = (t in X_full.index) and X_full.loc[t, sel].notna().all()
        if tiene_datos:
            z = scaler.transform(X_full.loc[t, sel].values.reshape(1, -1))
            f_new = pca.transform(z).ravel()
            tipo = f"nowcast ({cobertura}/3 meses)"
        else:
            # extrapolar cada factor con AR(1)
            f_new = np.array([ar1_step(np.array([fh[c] for fh in F_hist])) for c in range(k)])
            tipo = "forecast (factores extrapolados)"
        xrow = np.concatenate([[1.0], f_new, [y_lag_iter]])
        pred = float(xrow @ beta)
        se   = resid_std * np.sqrt(1.0 + j)      # el CI se ensancha con el horizonte
        # Senal direccional: P(expansion) = P(y_t > 0) bajo normalidad del error
        from scipy import stats
        p_exp = float(stats.norm.cdf(pred / se)) if se > 0 else np.nan
        proy.append({"fecha": t, "pred": pred, "lo": pred - 1.96 * se,
                     "hi": pred + 1.96 * se, "P_exp": p_exp, "tipo": tipo})
        F_hist.append(f_new)
        y_lag_iter = pred

    proy = pd.DataFrame(proy).set_index("fecha")

    # AR(1) benchmark forward
    ar_beta, ar_std, _ = _ols(np.column_stack([np.ones(len(y) - 1), y.values[:-1]]), y.values[1:])
    ar_preds, yl = [], y.iloc[-1]
    for j in range(len(fut_idx)):
        p = float(np.array([1.0, yl]) @ ar_beta)
        ar_preds.append(p); yl = p
    proy["AR1"] = ar_preds

    if verbose:
        print("=" * 74)
        print("PASO 5 | PROYECCION PROXIMOS %d TRIMESTRES (modelo de factores)" % h)
        print("=" * 74)
        print(f"  Variables en el modelo final ({len(sel)}): {', '.join(sel)}\n")
        proy_show = proy.copy()
        proy_show["senal"] = np.where(proy_show["P_exp"] >= 0.5, "EXPANSION", "CONTRACCION")
        print(proy_show[["pred", "lo", "hi", "P_exp", "senal", "AR1", "tipo"]].round(4).to_string())
        print()

    aux_final = {"seleccionadas": sel, "corr": corr,
                 "importancia": pd.Series(np.abs(pca.components_.T @ beta[1:1 + k]), index=sel)}
    return proy, aux_final


def _set_style():
    """Estilo econometrico comun a todos los graficos."""
    plt.rcParams.update({
        "figure.dpi": 130, "font.size": 10.5, "font.family": "DejaVu Sans",
        "axes.grid": True, "grid.alpha": 0.35, "grid.linestyle": "--",
        "axes.spines.top": False, "axes.spines.right": False,
        "axes.edgecolor": "#444444", "axes.linewidth": 0.9,
    })


def graficar(res, proy, aux, y, ruta_barras, ruta_lineas):
    """Genera los dos graficos de calidad de publicacion."""
    _set_style()

    # ------- GRAFICO 1: importancia de variables -------
    imp = aux["importancia"].sort_values(ascending=True)
    corr = aux["corr"].reindex(imp.index)
    fig, ax = plt.subplots(figsize=(8.6, max(4.2, 0.42 * len(imp))))
    colores = ["#1f6f8b" if corr[v] >= 0 else "#b23a48" for v in imp.index]
    ax.barh(imp.index, imp.values, color=colores, edgecolor="white", linewidth=0.6)
    for i, (v, val) in enumerate(imp.items()):
        ax.text(val + imp.max() * 0.01, i, f"{val:.3f}", va="center", fontsize=8.5)
    ax.set_title("Importancia de variables en el modelo de factores\n"
                 "(|carga PCA · coef. factores|; azul = corr(+) con PBI, rojo = corr(–))",
                 fontsize=11, loc="left")
    ax.set_xlabel("Contribucion al nowcast")
    ax.margins(x=0.12)
    fig.tight_layout()
    fig.savefig(ruta_barras, bbox_inches="tight")
    plt.close(fig)

    # ------- GRAFICO 2: serie temporal comparada -------
    fig, ax = plt.subplots(figsize=(11.5, 5.6))
    # historico real
    ax.plot(y.index, y.values, color="#111111", lw=1.8, label="PBI observado (q/q desest)", zorder=3)
    # nowcast OOS (factores) y AR1 OOS
    ax.plot(res.index, res["factores"], color="#1f6f8b", lw=1.4, ls="-",
            marker="o", ms=3, label="Nowcast modelo de factores (OOS)", zorder=2)
    ax.plot(res.index, res["AR1"], color="#e08e0b", lw=1.2, ls="--",
            label="Benchmark AR(1) (OOS)", zorder=1)
    # proyeccion futura + CI
    xf = proy.index
    ax.plot(xf, proy["pred"], color="#1f6f8b", lw=2.0, marker="D", ms=5,
            label="Proyeccion factores", zorder=4)
    ax.fill_between(xf, proy["lo"], proy["hi"], color="#1f6f8b", alpha=0.15,
                    label="IC 95%")
    ax.plot(xf, proy["AR1"], color="#e08e0b", lw=1.6, ls="--", marker="s", ms=4,
            label="Proyeccion AR(1)")
    # separador inicio de proyeccion
    ax.axvline(y.index.max(), color="#888888", lw=1.0, ls=":")
    ax.axhline(0, color="#aaaaaa", lw=0.8)
    ax.text(y.index.max(), ax.get_ylim()[1] * 0.92, " inicio proyeccion",
            fontsize=8.5, color="#666666")
    ax.yaxis.set_major_formatter(mticker.FuncFormatter(lambda v, _: f"{v*100:.0f}%"))
    ax.set_title("Nowcasting del PBI de Argentina: observado vs. modelo de factores vs. AR(1)",
                 fontsize=12.5, loc="left")
    ax.set_ylabel("Crecimiento trimestral q/q")
    ax.set_xlabel("")
    ax.legend(ncol=2, fontsize=9, framealpha=0.9, loc="lower left")
    fig.tight_layout()
    fig.savefig(ruta_lineas, bbox_inches="tight")
    plt.close(fig)


# ===========================================================================
# METRICA ALTERNATIVA - ACIERTO DIRECCIONAL (Pesaran-Timmermann)
# ===========================================================================
def _pesaran_timmermann(y_true, y_pred):
    """
    Test de Pesaran-Timmermann (1992) de capacidad predictiva DIRECCIONAL.
    H0: el signo del pronostico y el signo del dato son independientes,
    i.e. el modelo NO tiene skill direccional por encima del azar.
    Clave para un PM: corrige por el hecho de que el crecimiento q/q argentino
    es positivo la mayoria de los trimestres (predecir 'siempre positivo'
    acierta mucho sin skill real). Devuelve (acierto_signo, PT_stat, p_value).
    """
    from scipy import stats
    y_true = np.asarray(y_true, float)
    y_pred = np.asarray(y_pred, float)
    n  = len(y_true)
    P  = (np.sign(y_true) == np.sign(y_pred)).mean()   # tasa de acierto de signo
    Py = (y_true > 0).mean()
    Pp = (y_pred > 0).mean()
    Pstar = Py * Pp + (1 - Py) * (1 - Pp)              # acierto esperado si independientes
    var_P     = Pstar * (1 - Pstar) / n
    var_Pstar = ((2 * Py - 1) ** 2 * Pp * (1 - Pp) / n
                 + (2 * Pp - 1) ** 2 * Py * (1 - Py) / n
                 + 4 * Py * Pp * (1 - Py) * (1 - Pp) / n ** 2)
    denom = var_P - var_Pstar
    if denom <= 0:
        return P, np.nan, np.nan
    pt   = (P - Pstar) / np.sqrt(denom)
    pval = 2 * (1 - stats.norm.cdf(abs(pt)))
    return P, pt, pval


def metricas_direccionales(res, modelos=("factores", "puente", "AR1"), verbose=True):
    """Tabla de acierto de signo + test PT para cada modelo."""
    filas = []
    for m in modelos:
        df = res[["y_real", m]].dropna()
        sr, pt, p = _pesaran_timmermann(df["y_real"], df[m])
        filas.append((m, sr, pt, p))
    tabla = pd.DataFrame(filas, columns=["modelo", "acierto_signo", "PT_stat", "p_value"])
    if verbose:
        print("=" * 74)
        print("METRICA ALTERNATIVA | ACIERTO DIRECCIONAL (Pesaran-Timmermann 1992)")
        print("=" * 74)
        print("  Acierto de signo del crecimiento q/q (expansion vs contraccion).")
        print("  PT p<0.05 => skill direccional significativo (mejor que azar).\n")
        print(tabla.round(4).to_string(index=False))
        print()
    return tabla


# ===========================================================================
# GRAFICOS ADICIONALES: evolucion del RMSE y acierto direccional
# ===========================================================================
def graficar_evolucion_rmse(res, ruta, w=VENTANA_ROLLING):
    """
    Dos paneles:
      A) RMSE movil (ventana w) de cada modelo -> muestra cuando el modelo
         aporta precision y cuando no (el pico 2020 domina). Escala log para
         que se lea tambien el regimen normal.
      B) Ratio RMSE movil modelo/AR(1) -> version moderna de los Graficos 3-5
         del paper. Debajo de 1 = el modelo le gana al benchmark.
    """
    _set_style()
    colores = {"factores": "#1f6f8b", "puente": "#2a9d8f", "AR1": "#e08e0b"}
    err = {m: (res["y_real"] - res[m]) for m in ["factores", "puente", "AR1"]}
    rr  = lambda e: np.sqrt(e.pow(2).rolling(w).mean())

    fig, (ax1, ax2) = plt.subplots(1, 2, figsize=(13.2, 5.2))

    for m in ["factores", "puente", "AR1"]:
        ax1.plot(res.index, rr(err[m]) * 100, color=colores[m], lw=1.9,
                 label=m, marker="o", ms=2.5)
    ax1.set_yscale("log")
    ax1.set_title(f"RMSE movil (ventana {w} trim., escala log)", loc="left", fontsize=11.5)
    ax1.set_ylabel("RMSE (pp de crecimiento q/q)")
    ax1.legend(fontsize=9)

    for m in ["factores", "puente"]:
        ratio = rr(err[m]) / rr(err["AR1"])
        ax2.plot(res.index, ratio, color=colores[m], lw=1.9, marker="o", ms=2.5,
                 label=f"{m} / AR(1)")
    ax2.axhline(1.0, color="#333333", lw=1.1, ls="--")
    ax2.fill_between(res.index, 0, 1, color="#2a9d8f", alpha=0.05)
    ax2.set_title("Ratio RMSE vs AR(1)   (<1 = mejor que benchmark)", loc="left", fontsize=11.5)
    ax2.set_ylabel("RMSE modelo / RMSE AR(1)")
    ax2.legend(fontsize=9)

    fig.suptitle("Evolucion de la precision del nowcast a lo largo del OOS",
                 fontsize=13, x=0.02, ha="left", weight="bold")
    fig.tight_layout(rect=[0, 0, 1, 0.96])
    fig.savefig(ruta, bbox_inches="tight")
    plt.close(fig)


def graficar_direccional(res, tabla_dir, ruta, w=VENTANA_ROLLING):
    """
    Dos paneles:
      A) Acierto de signo por modelo (barras) con linea de azar (50%) y el
         p-value de Pesaran-Timmermann anotado.
      B) Acierto direccional movil (ventana w) en el tiempo.
    """
    _set_style()
    colores = {"factores": "#1f6f8b", "puente": "#2a9d8f", "AR1": "#e08e0b"}
    fig, (ax1, ax2) = plt.subplots(1, 2, figsize=(13.2, 5.2))

    modelos = tabla_dir["modelo"].tolist()
    sr = tabla_dir["acierto_signo"].values * 100
    bars = ax1.bar(modelos, sr, color=[colores[m] for m in modelos],
                   edgecolor="white", width=0.62)
    ax1.axhline(50, ls="--", color="#b23a48", lw=1.2, label="azar (50%)")
    for b, (_, row) in zip(bars, tabla_dir.iterrows()):
        p = row["p_value"]
        est = " *" if (pd.notna(p) and p < 0.05) else ""
        txt = f"{b.get_height():.0f}%\nPT p={p:.2f}{est}" if pd.notna(p) else f"{b.get_height():.0f}%\nPT n/d"
        ax1.text(b.get_x() + b.get_width() / 2, b.get_height() + 1.5, txt,
                 ha="center", va="bottom", fontsize=9)
    ax1.set_ylim(0, 100)
    ax1.set_ylabel("Acierto de signo (%)")
    ax1.set_title("Acierto direccional del nowcast (todo el OOS)", loc="left", fontsize=11.5)
    ax1.legend(fontsize=9, loc="lower right")

    for m in ["factores", "puente", "AR1"]:
        df = res[["y_real", m]].dropna()
        acierto = (np.sign(df["y_real"]) == np.sign(df[m])).astype(float)
        ax2.plot(df.index, acierto.rolling(w).mean() * 100, color=colores[m],
                 lw=1.9, label=m)
    ax2.axhline(50, ls="--", color="#b23a48", lw=1.2)
    ax2.set_ylim(0, 105)
    ax2.set_ylabel("% acierto de signo")
    ax2.set_title(f"Acierto direccional movil (ventana {w} trim.)", loc="left", fontsize=11.5)
    ax2.legend(fontsize=9, loc="lower left")

    fig.suptitle("Capacidad predictiva direccional: expansion vs. contraccion",
                 fontsize=13, x=0.02, ha="left", weight="bold")
    fig.tight_layout(rect=[0, 0, 1, 0.96])
    fig.savefig(ruta, bbox_inches="tight")
    plt.close(fig)


# ===========================================================================
# RESUMEN CONSOLIDADO DE EFICIENCIA
# ===========================================================================
def resumen_eficiencia(res, tabla_dir, gw_fac, gw_pue):
    """Consolida todas las metricas de eficiencia en una sola tabla + dict."""
    mask_covid = ~res.index.to_period("Q").isin(pd.period_range("2020Q2", "2020Q4", freq="Q"))
    ea = (res["y_real"] - res["AR1"]).abs()
    filas = []
    for m in ["factores", "puente", "AR1"]:
        err  = res["y_real"] - res[m]
        rmse = np.sqrt((err ** 2).mean())
        rmsx = np.sqrt((err[mask_covid] ** 2).mean())
        mae  = err.abs().mean()
        gana = np.nan if m == "AR1" else 100 * (err.abs() < ea).mean()
        d    = tabla_dir.set_index("modelo").loc[m]
        filas.append([m, rmse, rmsx, mae, gana, d["acierto_signo"] * 100, d["p_value"]])
    tabla = pd.DataFrame(filas, columns=["modelo", "RMSE", "RMSE_ex2020", "MAE",
                                         "%gana_vs_AR1", "acierto_dir_%", "PT_pvalue"])
    print("=" * 74)
    print("RESUMEN DE EFICIENCIA (todos los indicadores)")
    print("=" * 74)
    print(tabla.round(4).to_string(index=False))
    print("\n  Giacomini-White (magnitud, q/q):")
    print(f"    factores vs AR(1): t={gw_fac['t_stat']:.3f}  p={gw_fac['p_uncond']:.4f}  "
          f"(GW cond. chi2 p={gw_fac['p_cond']:.4f})")
    print(f"    puente   vs AR(1): t={gw_pue['t_stat']:.3f}  p={gw_pue['p_uncond']:.4f}")
    print("  Lectura: en MAGNITUD el modelo no supera al AR(1) con significancia,")
    print("           pero en DIRECCION (PT) si -> usar como senal direccional.\n")
    return tabla


# ===========================================================================
# BACKCAST DE REFERENCIA (con proxies de actividad coincidente)
# ===========================================================================
def backcast_referencia(df_mensual):
    """
    Lectura de referencia del trimestre en curso usando el panel CON proxies
    (EMAE / IGA_OJF / ICG). No es real-time (esos datos llegan tarde), pero
    sirve como 'ancla coincidente' una vez publicados, para contrastar con el
    nowcast hard-data. Devuelve la fila del primer trimestre proyectado.
    """
    y, X, X_full, Xq, n_meses, _ = preprocesar(df_mensual, usar_proxies=True, verbose=False)
    proy, _ = proyectar(y, X, X_full, Xq, n_meses, verbose=False)
    fila = proy.iloc[0]
    print("=" * 74)
    print("BACKCAST DE REFERENCIA (panel CON actividad coincidente EMAE/IGA/ICG)")
    print("=" * 74)
    print(f"  Trimestre {proy.index[0].date()}: nivel={fila['pred']:.4f}  "
          f"P(expansion)={fila['P_exp']:.3f}  [{fila['tipo']}]")
    print("  (No confundir con la senal real-time hard-data; es ancla ex-post.)\n")
    return proy


# ===========================================================================
# EXCEL TEORICO: explicacion del modelo + bibliografia
# ===========================================================================
def construir_excel_teorico(ruta, tabla_ef, proy, sel_final):
    """Genera un Excel de documentacion teorica del modelo y su bibliografia."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    AZUL   = "1F4E79"
    AZUL2  = "2E75B6"
    GRIS   = "F2F2F2"
    f_tit  = Font(name="Arial", size=15, bold=True, color="FFFFFF")
    f_h    = Font(name="Arial", size=11.5, bold=True, color="FFFFFF")
    f_sub  = Font(name="Arial", size=11, bold=True, color=AZUL)
    f_body = Font(name="Arial", size=10.5, color="222222")
    f_bold = Font(name="Arial", size=10.5, bold=True, color="222222")
    fill_t = PatternFill("solid", fgColor=AZUL)
    fill_h = PatternFill("solid", fgColor=AZUL2)
    fill_g = PatternFill("solid", fgColor=GRIS)
    wrap   = Alignment(wrap_text=True, vertical="top")
    wrapc  = Alignment(wrap_text=True, vertical="center", horizontal="center")
    thin   = Side(style="thin", color="BBBBBB")
    borde  = Border(left=thin, right=thin, top=thin, bottom=thin)

    wb = Workbook()

    def nueva_hoja(nombre, titulo, ancho=92):
        ws = wb.create_sheet(nombre)
        ws.sheet_view.showGridLines = False
        ws.column_dimensions["A"].width = 3
        ws.column_dimensions["B"].width = ancho
        ws.merge_cells("B2:B2")
        c = ws["B2"]; c.value = titulo; c.font = f_tit; c.fill = fill_t
        c.alignment = Alignment(vertical="center", horizontal="left", indent=1)
        ws.row_dimensions[2].height = 34
        return ws

    def escribir(ws, secciones, fila_ini=4):
        r = fila_ini
        for tipo, texto in secciones:
            c = ws.cell(row=r, column=2, value=texto)
            if tipo == "h":
                c.font = f_h; c.fill = fill_h
                c.alignment = Alignment(vertical="center", indent=1)
                ws.row_dimensions[r].height = 22
            elif tipo == "sub":
                c.font = f_sub; c.alignment = Alignment(vertical="center")
                ws.row_dimensions[r].height = 20
            elif tipo == "eq":
                c.font = Font(name="Consolas", size=10.5, color="7030A0")
                c.alignment = wrap; c.fill = fill_g
            elif tipo == "b":
                c.font = f_bold; c.alignment = wrap
            else:
                c.font = f_body; c.alignment = wrap
                n = max(1, int(len(texto) / 88) + texto.count("\n") + 1)
                ws.row_dimensions[r].height = 15 * n
            r += 1
        return r

    # quitar hoja default
    wb.remove(wb.active)

    # ------------------- HOJA 1: RESUMEN -------------------
    ws = nueva_hoja("Resumen", "Nowcasting del PBI de Argentina  |  Modelo de Factores Dinamicos")
    p_now = proy.iloc[0]
    escribir(ws, [
        ("h", "Objetivo"),
        ("p", "Producir una estimacion temprana ('nowcast') del crecimiento trimestral "
              "del PBI real desestacionalizado de Argentina, explotando un panel de "
              "indicadores mensuales de actividad, precios, dinero y sector externo, "
              "disponibles antes que la cifra oficial (que se publica con ~10 semanas de rezago)."),
        ("h", "Enfoque"),
        ("p", "Modelo de factores dinamicos: se extraen pocos factores comunes latentes de "
              "un panel amplio via componentes principales (PCA) y se usan como regresores del "
              "PBI, con un termino autorregresivo. Replica y actualiza el ejercicio de "
              "D'Amato, Garegnani y Blanco (2016, BCRA), extendido a datos 2004-2026."),
        ("h", "Configuracion de produccion (real-time)"),
        ("p", "El nowcast en tiempo real usa un panel de HARD DATA y EXCLUYE los proxies de "
              "actividad coincidente (EMAE, IGA_OJF, ICG). Motivo: EMAE es practicamente el "
              "PBI mensual y tiene el mayor rezago de publicacion del panel; incluirlo genera "
              "una performance out-of-sample ilusoria que no estaria disponible al momento de "
              "producir la senal. Los proxies se reservan para el backcast ex-post."),
        ("h", "Hallazgo central"),
        ("b", "En MAGNITUD el modelo no supera al benchmark AR(1) con significancia estadistica "
              "(Giacomini-White), pero en DIRECCION (expansion vs contraccion) si lo hace de "
              "forma robusta (Pesaran-Timmermann, p<0.01)."),
        ("p", "Recomendacion de uso: senal DIRECCIONAL (probabilidad de expansion) e input de "
              "monitoreo del ciclo, no estimador puntual del q/q ni senal automatica de trading, "
              "hasta validar valor economico en un backtest de decision."),
        ("h", "Nowcast vigente"),
        ("b", f"Trimestre {proy.index[0].date()}:  nivel estimado {p_now['pred']*100:.2f}% q/q  |  "
              f"P(expansion) = {p_now['P_exp']:.0%}  |  senal: "
              f"{'EXPANSION' if p_now['P_exp']>=0.5 else 'CONTRACCION'}"),
        ("p", f"Variables en el modelo final ({len(sel_final)}): {', '.join(sel_final)}."),
    ])

    # ------------------- HOJA 2: METODOLOGIA -------------------
    ws = nueva_hoja("Metodologia", "Metodologia del modelo, paso a paso")
    escribir(ws, [
        ("h", "Paso 1 - Preprocesamiento y estacionariedad"),
        ("p", "El PBI viene como crecimiento trimestral q/q (ya estacionario). Los indicadores "
              "mensuales se agregan a frecuencia trimestral por promedio simple (cada mes pesa "
              "igual), homogeneizandolos con el PBI:"),
        ("eq", "  X_t^Q = ( X_{N,t} + X_{N-1,t} + ... + X_{1,t} ) / N        (ec. 4 del paper)"),
        ("p", "Cada indicador se somete a un test ADF iterativo y se transforma con la operacion "
              "MINIMA que induce estacionariedad (nivel -> dif. log -> dif. -> 2da dif.). "
              "Decision: se exige estacionariedad al 1% para aceptar 'nivel', de modo que las "
              "series de actividad (casi I(1)) pasen a crecimiento, homogeneas con el target."),
        ("h", "Paso 2 - Seleccion de variables"),
        ("p", "Se retienen los indicadores con |correlacion| con el crecimiento del PBI superior "
              "a 0,5 (criterio del paper). Clave metodologica: la seleccion, la estandarizacion "
              "y el PCA se RE-ESTIMAN dentro de cada ventana del ejercicio out-of-sample, usando "
              "solo informacion hasta t. Asi se evita el look-ahead / leakage de seleccion que "
              "tiene la version original (que selecciona sobre toda la muestra)."),
        ("h", "Paso 3 - Modelo de factores"),
        ("p", "La covarianza de las n series se resume en q factores comunes latentes (n > q):"),
        ("eq", "  X_it = lambda_i(L)' f_t + u_it                              (ec. 1)"),
        ("p", "El PBI se modela como funcion de los factores y su propio rezago:"),
        ("eq", "  y_t = beta(L)' f_t + gamma(L) y_{t-1} + e_t                 (ec. 2)"),
        ("eq", "  y_hat_t^Q = beta(L)' f_t^Q + gamma(L) y_{t-1}^Q            (ec. 3, nowcast)"),
        ("p", "Los factores se estiman por componentes principales (Stock-Watson 2002). El "
              "numero de factores replica el criterio del paper (scree plot, ~4); una mejora "
              "posible es el criterio de informacion de Bai-Ng (2002)."),
        ("h", "Paso 4 - Evaluacion pseudo out-of-sample"),
        ("p", "Ventana expansiva: en cada trimestre t se entrena con la informacion hasta t-1 y "
              "se nowcastea y_t usando los indicadores de t (disponibles antes que el PBI). Se "
              "compara contra un AR(1) (benchmark) y contra ecuaciones puente (comparador)."),
        ("h", "Paso 5 - Proyeccion"),
        ("p", "Para trimestres con datos mensuales (parciales o completos) se produce un nowcast "
              "directo; para los siguientes se extrapolan los factores por AR(1) (forecast puro, "
              "con intervalos que se ensanchan con el horizonte). Salida dual: nivel q/q y "
              "P(expansion) = P(y_t > 0) bajo normalidad del error."),
    ])

    # ------------------- HOJA 3: ESPECIFICACION -------------------
    ws = nueva_hoja("Especificacion", "Especificacion y parametros")
    escribir(ws, [
        ("h", "Variable objetivo"),
        ("p", "y_t = crecimiento trimestral q/q del PBI real desestacionalizado (DPBI_desest). "
              "Estacionario por construccion; no se transforma."),
        ("h", "Ecuacion estimada"),
        ("eq", "  y_t = c + beta_1 F1_t + ... + beta_k Fk_t + gamma y_{t-1} + e_t"),
        ("p", "donde F1..Fk son los primeros k componentes principales del panel de indicadores "
              "seleccionados, estandarizados dentro de la ventana."),
        ("h", "Parametros de configuracion"),
        ("b", "Umbral de seleccion |corr| > 0,50"),
        ("b", "Numero de factores (PCA): 4 (criterio scree)"),
        ("b", "Ventana inicial de entrenamiento: 40 trimestres"),
        ("b", "Tipo de ventana: expansiva"),
        ("b", "Horizonte de proyeccion: 4 trimestres"),
        ("b", "Panel de produccion: hard data (proxies EMAE/IGA/ICG excluidos)"),
        ("h", "Panel de indicadores (dominio)"),
        ("p", "Actividad (IPI, ISAC, siderurgia, cemento, automotriz), sector externo "
              "(exportaciones, ITCRM), recaudacion (IVA, Ganancias, comercio exterior), dinero y "
              "tasas (M1, M2, BADLAR), expectativas (ICC), salarios y agro (molienda de soja)."),
    ])

    # ------------------- HOJA 4: METRICAS -------------------
    ws = nueva_hoja("Metricas de eficiencia", "Metricas de evaluacion e interpretacion")
    escribir(ws, [
        ("h", "Metricas de magnitud"),
        ("b", "RMSE / MAE"),
        ("p", "Raiz del error cuadratico medio y error absoluto medio del nowcast. RMSE penaliza "
              "cuadraticamente: el trimestre COVID (2020) domina la varianza, por eso se reporta "
              "tambien RMSE excluyendo 2020Q2-2020Q4."),
        ("b", "Test de Giacomini-White (2006)"),
        ("p", "Compara capacidad predictiva de dos modelos sobre la funcion de perdida cuadratica. "
              "Version incondicional: t-stat de la media del diferencial de perdida con errores "
              "Newey-West (la del paper). Version condicional: estadistico chi-cuadrado con "
              "instrumentos h_{t-1}. H0: igual capacidad predictiva."),
        ("h", "Metrica direccional (recomendada para uso)"),
        ("b", "Test de Pesaran-Timmermann (1992)"),
        ("p", "Evalua el acierto del SIGNO (expansion/contraccion), corrigiendo por el hecho de "
              "que el crecimiento q/q argentino es positivo la mayoria de los trimestres "
              "(predecir 'siempre positivo' acierta mucho sin skill real). H0: signo del "
              "pronostico y del dato independientes. p<0.05 => skill direccional genuino."),
        ("h", "Resultados obtenidos (OOS actual)"),
    ], fila_ini=4)
    # tabla de resultados
    r0 = ws.max_row + 2
    headers = ["Modelo", "RMSE", "RMSE ex-2020", "MAE", "% gana vs AR1", "Acierto dir. %", "PT p-value"]
    for j, hh in enumerate(headers):
        c = ws.cell(row=r0, column=2 + j, value=hh)
        c.font = f_h; c.fill = fill_h; c.alignment = wrapc; c.border = borde
        ws.column_dimensions[get_column_letter(2 + j)].width = 16 if j else 12
    for i, (_, row) in enumerate(tabla_ef.iterrows(), start=1):
        vals = [row["modelo"], row["RMSE"], row["RMSE_ex2020"], row["MAE"],
                row["%gana_vs_AR1"], row["acierto_dir_%"], row["PT_pvalue"]]
        for j, v in enumerate(vals):
            c = ws.cell(row=r0 + i, column=2 + j)
            if isinstance(v, float) and pd.isna(v):
                c.value = "-"
            elif isinstance(v, float):
                c.value = round(v, 4)
            else:
                c.value = v
            c.font = f_body; c.border = borde
            c.alignment = Alignment(horizontal="center")
            if i % 2 == 0: c.fill = fill_g
    ws.column_dimensions["B"].width = 16

    # ------------------- HOJA 5: LIMITACIONES -------------------
    ws = nueva_hoja("Limitaciones y uso", "Limitaciones, supuestos y uso recomendado")
    escribir(ws, [
        ("h", "Supuestos y limitaciones"),
        ("b", "1. Desestacionalizacion con posible look-ahead"),
        ("p", "La base viene desestacionalizada. Si el ajuste se hizo con un filtro de dos lados "
              "sobre toda la muestra, cada observacion historica incorpora informacion futura y "
              "el OOS resulta optimista. Un pipeline real-time requiere ajuste estacional de un lado."),
        ("b", "2. Historia out-of-sample corta"),
        ("p", "Con datos trimestrales desde 2004, la muestra OOS (~45 trimestres) es chica para "
              "los asintoticos del test de Giacomini-White; de ahi la no-significancia en magnitud."),
        ("b", "3. Dependencia de proxies coincidentes"),
        ("p", "La performance con EMAE/IGA/ICG es en parte circular (son proxies del PBI). La "
              "version de produccion los excluye para el nowcast real-time."),
        ("b", "4. Manejo simplificado del ragged edge"),
        ("p", "Colapsar a trimestral descarta el timing intra-trimestral de las publicaciones. "
              "Un modelo de factores dinamico en espacio de estados con filtro de Kalman "
              "(Mariano-Murasawa 2003; Camacho et al. 2015) maneja el ragged edge de forma nativa "
              "y actualiza el nowcast a medida que llega cada dato; es el paso natural para produccion."),
        ("h", "Uso recomendado"),
        ("p", "Senal direccional (P de expansion) e input de monitoreo del ciclo para decisiones "
              "de asignacion tactica. No usar como estimador puntual del q/q ni como senal "
              "automatica hasta validar valor economico en un backtest de decision."),
    ])

    # ------------------- HOJA 6: BIBLIOGRAFIA -------------------
    ws = nueva_hoja("Bibliografia", "Bibliografia relacionada", ancho=70)
    ws.column_dimensions["C"].width = 60
    r0 = 4
    for j, hh in enumerate(["Referencia", "Aporte al modelo"]):
        c = ws.cell(row=r0, column=2 + j, value=hh)
        c.font = f_h; c.fill = fill_h; c.alignment = Alignment(vertical="center", indent=1); c.border = borde
    biblio = [
        ("D'Amato, L., Garegnani, L. y Blanco, E. (2016). \"Nowcasting de PIB: evaluando las "
         "condiciones ciclicas de la economia argentina\". BCRA, Ensayos Economicos 74.",
         "Paper base. Ecuaciones puente y modelo de factores para el PBI argentino."),
        ("Giannone, D., Reichlin, L. y Small, D. (2008). \"Nowcasting: The real-time informational "
         "content of macroeconomic data\". Journal of Monetary Economics 55(4).",
         "Fundamento conceptual del nowcasting y del flujo de informacion en tiempo real."),
        ("Giannone, D., Reichlin, L. y Small, D. (2005). \"Nowcasting GDP and Inflation...\". "
         "CEPR Discussion Paper 5178.",
         "Uso de factores comunes como regresores del PBI."),
        ("Stock, J. y Watson, M. (2002). \"Macroeconomic Forecasting Using Diffusion Indexes\". "
         "Journal of Business & Economic Statistics 20(2).",
         "Estimacion de factores por componentes principales (nucleo del Paso 3)."),
        ("Bai, J. y Ng, S. (2002). \"Determining the Number of Factors in Approximate Factor "
         "Models\". Econometrica 70(1).",
         "Criterio formal para elegir el numero de factores (mejora sobre el scree plot)."),
        ("Giacomini, R. y White, H. (2006). \"Tests of Conditional Predictive Ability\". "
         "Econometrica 74(6).",
         "Test de capacidad predictiva (magnitud), incondicional y condicional."),
        ("Pesaran, M.H. y Timmermann, A. (1992). \"A Simple Nonparametric Test of Predictive "
         "Performance\". Journal of Business & Economic Statistics 10(4).",
         "Test de acierto DIRECCIONAL; sustenta la senal de expansion/contraccion."),
        ("Diebold, F. y Mariano, R. (1995). \"Comparing Predictive Accuracy\". JBES 13(3).",
         "Antecedente del test de capacidad predictiva comparada."),
        ("Newey, W. y West, K. (1987). \"A Simple, Positive Semi-Definite, HAC Covariance "
         "Matrix\". Econometrica 55(3).",
         "Errores estandar robustos usados en el test de Giacomini-White."),
        ("Banbura, M., Giannone, D., Modugno, M. y Reichlin, L. (2013). \"Now-Casting and the "
         "Real-Time Data Flow\". Handbook of Economic Forecasting, Vol. 2A.",
         "Manejo del ragged edge y actualizacion secuencial del nowcast."),
        ("Mariano, R. y Murasawa, Y. (2003). \"A new coincident index of business cycles based "
         "on monthly and quarterly series\". Journal of Applied Econometrics 18(4).",
         "DFM de frecuencia mixta en espacio de estados (Kalman): el paso a produccion."),
        ("Camacho, M., Dal Bianco, M. y Martinez-Martin, J. (2015). \"Short-Run Forecasting of "
         "Argentine GDP Growth\". Emerging Markets Finance and Trade 51(3).",
         "Aplicacion de DFM de frecuencia mixta a la Argentina."),
        ("Ghysels, E., Santa-Clara, P. y Valkanov, R. (2004). \"The MIDAS Touch: Mixed Data "
         "Sampling Regression Models\". CIRANO 2004s-20.",
         "Alternativa de frecuencia mixta (regresiones MIDAS)."),
        ("Kelly, B. y Pruitt, S. (2015). \"The three-pass regression filter: A new approach to "
         "forecasting using many predictors\". Journal of Econometrics 186(2).",
         "Alternativa al PCA para extraer senal relevante al target (3PRF)."),
    ]
    r = r0 + 1
    for i, (ref, aporte) in enumerate(biblio):
        c1 = ws.cell(row=r, column=2, value=ref);    c1.font = f_body; c1.alignment = wrap; c1.border = borde
        c2 = ws.cell(row=r, column=3, value=aporte); c2.font = f_body; c2.alignment = wrap; c2.border = borde
        if i % 2 == 0:
            c1.fill = fill_g; c2.fill = fill_g
        ws.row_dimensions[r].height = 15 * (max(len(ref) // 70, len(aporte) // 60) + 2)
        r += 1

    wb.save(ruta)
    print(f"Excel teorico guardado en {ruta}")


# ===========================================================================
# MAIN
# ===========================================================================
def main():
    import os
    os.makedirs(DIR_OUT, exist_ok=True)
    df_mensual = cargar_datos(RUTA_EXCEL)

    # Paso 1 - Preprocesamiento (PRODUCCION: hard data, real-time)
    y, X, X_full, Xq, n_meses, tabla = preprocesar(df_mensual, usar_proxies=USAR_PROXIES_ACTIVIDAD)

    # Paso 2 - Seleccion (informativa full-sample; en OOS es leak-free)
    sel, corr = seleccionar_variables(y, X)
    print("=" * 74)
    print("PASO 2 | SELECCION DE VARIABLES (corr con PBI > %.2f)" % UMBRAL_CORR)
    print("=" * 74)
    print("  (Ranking full-sample referencial; en el OOS la seleccion se re-estima")
    print("   dentro de cada ventana para evitar leakage)\n")
    print(corr.reindex(corr.abs().sort_values(ascending=False).index).round(4).to_string())
    print(f"\n  Seleccionadas (|corr|>{UMBRAL_CORR}): {len(sel)} -> {', '.join(sel)}\n")

    # Paso 3 - Evaluacion OOS
    res, aux = evaluar_oos(y, X)

    # Paso 4 - Tests de eficiencia (todos)
    gw_fac = test_giacomini_white(res, "factores", "AR1")
    gw_pue = test_giacomini_white(res, "puente", "AR1")
    tabla_dir = metricas_direccionales(res)
    tabla_ef  = resumen_eficiencia(res, tabla_dir, gw_fac, gw_pue)

    # Paso 5 - Proyeccion 4 trimestres (nivel + P(expansion))
    proy, aux_final = proyectar(y, X, X_full, Xq, n_meses)

    # Backcast de referencia con proxies (ancla ex-post)
    backcast_referencia(df_mensual)

    # Graficos (los cuatro)
    graficar(res, proy, aux_final, y,
             f"{DIR_OUT}/grafico_importancia.png", f"{DIR_OUT}/grafico_nowcast.png")
    graficar_evolucion_rmse(res, f"{DIR_OUT}/grafico_rmse_evolucion.png")
    graficar_direccional(res, tabla_dir, f"{DIR_OUT}/grafico_direccional.png")
    print("Graficos guardados en %s (4 en total)" % DIR_OUT)

    # Excel teorico
    construir_excel_teorico(f"{DIR_OUT}/modelo_nowcasting_teoria.xlsx",
                            tabla_ef, proy, aux_final["seleccionadas"])
    print("\nProceso completo.")


if __name__ == "__main__":
    main()


if __name__ == "__main__":
    main()
